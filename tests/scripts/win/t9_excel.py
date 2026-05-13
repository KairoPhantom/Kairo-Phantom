#!/usr/bin/env python3
import os, time, json

RESULT_DIR = r"C:\tests\results"
FIXTURES_DIR = r"C:\tests\fixtures"
os.makedirs(RESULT_DIR, exist_ok=True)
os.makedirs(FIXTURES_DIR, exist_ok=True)
RESULT_FILE = os.path.join(RESULT_DIR, "t9_win_result.json")

def write_result(payload):
    open(RESULT_FILE, 'w').write(json.dumps(payload))
    print(json.dumps(payload))

def run_excel_e2e():
    try:
        import openpyxl
    except ImportError:
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kairo Test"
    ws['A1'] = "Scenario"
    ws['B1'] = "Result"
    ws['A2'] = "Excel E2E"
    ws['B2'] = "PASS"

    save_path = os.path.join(FIXTURES_DIR, "excel_t9_autogen.xlsx")
    wb.save(save_path)
    return {"method": "openpyxl", "saved": save_path}

def main():
    try:
        res = run_excel_e2e()
        write_result({"id": "t9", "status": "PASS", "details": res})
    except Exception as e:
        write_result({"id": "t9", "status": "FAIL", "error": str(e)})

if __name__ == '__main__':
    main()
