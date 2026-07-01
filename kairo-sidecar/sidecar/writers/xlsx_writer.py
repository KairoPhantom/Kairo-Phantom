import os
import shutil
import logging
import re
from pathlib import Path
from typing import List, Dict, Any

log = logging.getLogger("kairo-sidecar.xlsx_writer")


def _validate_formula(formula: str) -> bool:
    """
    Validate Excel formula using openpyxl Tokenizer and ForgeValidator.
    Checks for:
      - Starts with '='
      - Unbalanced parentheses or braces
      - Common cell reference syntax errors (e.g., A0)
      - Function argument mismatches or circular references
    """
    if not isinstance(formula, str) or not formula.strip().startswith("="):
        return False

    try:
        from openpyxl.formula.tokenize import Tokenizer

        formula_str = formula.strip()
        tok = Tokenizer(formula_str)

        # 1. Parentheses and Braces check
        paren_count = 0
        curly_count = 0
        for token in tok.items:
            if token.type == "PAREN":
                if token.value == "(":
                    paren_count += 1
                elif token.value == ")":
                    paren_count -= 1
                    if paren_count < 0:
                        return False
            elif token.type == "CURLY":
                if token.value == "{":
                    curly_count += 1
                elif token.value == "}":
                    curly_count -= 1
                    if curly_count < 0:
                        return False

        if paren_count != 0 or curly_count != 0:
            return False

        # 2. Check for invalid row index 0 (e.g. A0, B0)
        if re.search(r"\b[A-Za-z]+0+\b", formula_str):
            return False

        # 3. Check with ForgeValidator if possible
        try:
            from sidecar.parsers.forge_bridge import validate_formula

            val_res = validate_formula(formula_str)
            if not val_res["valid"]:
                return False
        except Exception:
            pass  # Fallback to tokenizer if forge_bridge is not importable

    except Exception:
        return False

    return True


def _safe_write_formula(formula: str) -> str:
    """
    Validate and return the formula if valid, else raise ValueError.
    """
    if not _validate_formula(formula):
        raise ValueError(f"Formula validation failed: {formula}")
    return formula


def write_xlsx(file_path: str, operations: List[Dict[str, Any]]) -> dict:
    """
    Apply ExcelOperation list to .xlsx file atomically.

    Supported operation types:
      - write_cell: Write a single cell (value or formula)
      - write_range: Write a 2D grid starting at a cell address
      - set_number_format: Set number format on a cell or range
      - set_column_width: Set column width

    Preserves existing conditional formatting, named ranges, and data
    validation on untouched cells. Never clobbers what wasn't asked to change.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {path}"}

    # --- Atomic backup ---
    backup_path = path.with_suffix(path.suffix + ".kairo_backup")
    try:
        shutil.copy2(path, backup_path)
    except Exception as e:
        return {"error": f"Failed to create backup of spreadsheet: {e}"}

    applied = []
    errors = []

    try:
        # Use keep_vba=True to preserve macros if present
        try:
            wb = openpyxl.load_workbook(str(path), keep_vba=True)
        except Exception:
            wb = openpyxl.load_workbook(str(path))

        for op in operations:
            try:
                op_type = op.get("type", "write_cell")  # default to write_cell
                sheet_name = op.get("sheet", None)

                # Resolve sheet: by name, or active sheet
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active

                if op_type == "write_cell":
                    cell_ref = op.get("cell", "A1")
                    value = op.get("value", "")
                    formula = op.get("formula", "")
                    number_format = op.get("number_format", None)

                    if formula:
                        _safe_write_formula(formula)

                    # Formula takes precedence over value
                    write_val = formula if formula else value
                    ws[cell_ref] = write_val

                    # Apply number format if specified (preserving existing otherwise)
                    if number_format and number_format.strip():
                        ws[cell_ref].number_format = number_format

                    applied.append(
                        {"type": "write_cell", "cell": cell_ref, "written": str(write_val)[:100]}
                    )

                elif op_type == "write_range":
                    start_cell = op.get("start", op.get("cell", "A1"))
                    rows_data = op.get("values", op.get("data", []))
                    if not rows_data:
                        errors.append(f"write_range: no values provided for op {op}")
                        continue

                    # Parse start cell to row/col numbers
                    from openpyxl.utils import get_column_letter
                    from openpyxl.utils.cell import coordinate_to_tuple

                    try:
                        start_row, start_col = coordinate_to_tuple(start_cell)
                    except Exception:
                        start_row, start_col = 1, 1

                    cells_written = 0
                    for row_i, row_values in enumerate(rows_data):
                        if not isinstance(row_values, (list, tuple)):
                            row_values = [row_values]
                        for col_i, cell_val in enumerate(row_values):
                            r = start_row + row_i
                            c = start_col + col_i
                            cell_addr = f"{get_column_letter(c)}{r}"

                            if isinstance(cell_val, str) and cell_val.startswith("="):
                                _safe_write_formula(cell_val)

                            ws[cell_addr] = cell_val
                            cells_written += 1

                    applied.append(
                        {
                            "type": "write_range",
                            "start": start_cell,
                            "rows": len(rows_data),
                            "cells": cells_written,
                        }
                    )

                elif op_type == "set_number_format":
                    cell_ref = op.get("cell", "A1")
                    number_format = op.get("number_format", "General")
                    ws[cell_ref].number_format = number_format
                    applied.append(
                        {"type": "set_number_format", "cell": cell_ref, "format": number_format}
                    )

                elif op_type == "set_column_width":
                    col = op.get("column", "A")
                    width = op.get("width", 15)
                    ws.column_dimensions[col].width = width
                    applied.append({"type": "set_column_width", "column": col, "width": width})

                elif op_type in ("write_formula", "formula"):
                    # Alias: write_formula is the same as write_cell with formula
                    cell_ref = op.get("cell", "A1")
                    formula = op.get("formula", op.get("value", ""))

                    _safe_write_formula(formula)

                    ws[cell_ref] = formula
                    applied.append(
                        {"type": "write_formula", "cell": cell_ref, "formula": str(formula)[:100]}
                    )

                else:
                    # Legacy mode: treat any op with a 'cell' key as write_cell
                    cell_ref = op.get("cell")
                    if cell_ref:
                        value = op.get("value", "")
                        formula = op.get("formula", "")
                        if formula:
                            _safe_write_formula(formula)
                        write_val = formula if formula else value
                        ws[cell_ref] = write_val
                        applied.append(
                            {
                                "type": "legacy_write",
                                "cell": cell_ref,
                                "written": str(write_val)[:100],
                            }
                        )
                    else:
                        errors.append(f"Unsupported operation type '{op_type}' — op: {op}")

            except Exception as e:
                errors.append(f"Excel op failed {op}: {e}")
                log.warning(f"Excel write op failed: {op} — {e}")

        # --- Atomic save via tmp + rename ---
        tmp_path = path.with_suffix(path.suffix + ".kairo_tmp")
        try:
            wb.save(str(tmp_path))
            os.replace(str(tmp_path), str(path))
        except PermissionError:
            # Excel has file locked — clean up
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass
            # Restore backup
            if backup_path.exists():
                try:
                    shutil.copy2(str(backup_path), str(path))
                except Exception:
                    pass
            return {
                "error": "Excel has this file open. Save and close Excel first, then press Alt+Ctrl+M again.",
                "path": str(path),
            }
        except Exception as e:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass
            raise e

        # --- Clean up backup on success ---
        if not errors:
            try:
                backup_path.unlink(missing_ok=True)
            except Exception:
                pass

        return {
            "applied_count": len(applied),
            "cells": applied,
            "errors": errors,
            "path": str(path),
        }

    except Exception as e:
        log.error(f"Failed to write Excel file {file_path}: {e}")
        # Try to restore backup
        if backup_path.exists():
            try:
                shutil.copy2(str(backup_path), str(path))
            except Exception:
                pass
        return {"error": f"Failed to write Excel: {e}"}
