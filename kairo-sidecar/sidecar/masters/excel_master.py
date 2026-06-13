import os
import shutil
import logging
import traceback
import re
from dataclasses import dataclass, asdict
from typing import List, Dict, Any, Union, Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

from sidecar.parsers.forge_bridge import ForgeValidator  # Single canonical source

log = logging.getLogger("kairo-sidecar.excel_master")


@dataclass
class ExcelContext:
    active_cell: str
    active_sheet: str
    sheet_names: List[str]
    cells: List[Dict[str, Any]]
    headers: Dict[str, str]
    named_ranges: Dict[str, str]
    column_types: Dict[str, str]
    locale: str
    max_row: int
    max_col: int
    file_path: str = ""
    tables: Optional[List[Dict[str, Any]]] = None

    def to_dict(self):
        return asdict(self)


class ValidationResult:
    def __init__(self, valid: bool, error: str = "", op: dict = None):
        self.valid = valid
        self.error = error
        self.op = op


class ExcelContextExtractor:
    """Reads spreadsheet context centered on active cell (15x15 region) with full details."""

    def extract(self, file_path: str, active_cell: str, active_sheet: Optional[str] = None) -> ExcelContext:
        try:
            wb = load_workbook(file_path, data_only=False)  # data_only=False → get formulas
            
            # Resolve active sheet
            if active_sheet and active_sheet in wb.sheetnames:
                ws = wb[active_sheet]
            else:
                ws = wb.active
                active_sheet = ws.title

            # Parse active_cell into row/col
            col_letter, row_num = self._parse_cell_address(active_cell)

            # Extract surrounding region (15 rows × 15 cols centered on active cell)
            min_row = max(1, row_num - 7)
            max_row = min(ws.max_row or 1, row_num + 7)
            min_col = max(1, col_letter - 7)
            max_col = min(ws.max_column or 1, col_letter + 7)

            cells = []
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cells.append({
                        "address": cell.coordinate,
                        "value": cell.value,
                        "formula": str(cell.value) if str(cell.value or "").startswith("=") else None,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "is_empty": cell.value is None,
                    })

            # Detect header row
            headers = self._detect_headers(ws)

            # Get named ranges
            named_ranges = {}
            try:
                if hasattr(wb.defined_names, "definedName"):
                    for nr in wb.defined_names.definedName:
                        named_ranges[nr.name] = str(nr.attr_text or nr.value or "")
                else:
                    for name, defn in wb.defined_names.items():
                        named_ranges[name] = str(defn.value or "")
            except Exception:
                pass

            # Detect column types
            column_types = self._infer_column_types(ws, headers)

            # Detect locale (check existing formulas for separator)
            locale = self._detect_locale(ws, active_cell)

            # Tables per sheet
            tables: List[Dict[str, Any]] = []
            try:
                for sname in wb.sheetnames:
                    ws2 = wb[sname]
                    for tname, tobj in (getattr(ws2, "tables", {}) or {}).items():
                        tables.append({
                            "name": tname,
                            "sheet": sname,
                            "range": getattr(tobj, "ref", "")
                        })
            except Exception:
                pass

            return ExcelContext(
                active_cell=active_cell,
                active_sheet=active_sheet,
                sheet_names=wb.sheetnames,
                cells=cells,
                headers=headers,
                named_ranges=named_ranges,
                column_types=column_types,
                locale=locale,  # "en" (comma) or "eu" (semicolon)
                max_row=ws.max_row or 0,
                max_col=ws.max_column or 0,
                file_path=file_path,
                tables=tables
            )
        except Exception as e:
            log.error(f"ExcelContextExtractor error: {traceback.format_exc()}")
            return ExcelContext(
                active_cell=active_cell,
                active_sheet=active_sheet or "Sheet1",
                sheet_names=[active_sheet or "Sheet1"],
                cells=[],
                headers={},
                named_ranges={},
                column_types={},
                locale="en",
                max_row=0,
                max_col=0,
                file_path=file_path,
                tables=[]
            )

    def _parse_cell_address(self, active_cell: str) -> tuple[int, int]:
        col_str = "".join(c for c in active_cell if c.isalpha())
        row_str = "".join(c for c in active_cell if c.isdigit())
        
        # Convert col_str to 1-indexed column number
        col_num = 0
        for char in col_str.upper():
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
        row_num = int(row_str) if row_str else 1
        return col_num if col_num > 0 else 1, row_num

    def _detect_headers(self, ws) -> Dict[str, str]:
        best_row = 1
        best_score = -1
        
        # Scan first 10 rows and at most first 100 columns
        max_scan = min(10, ws.max_row or 1)
        max_col = min(ws.max_column or 1, 100)
        
        # Use iter_rows to avoid individual ws.cell() lookups and creation
        rows = list(ws.iter_rows(min_row=1, max_row=max_scan, min_col=1, max_col=max_col, values_only=True))
        
        for r_idx, row in enumerate(rows, start=1):
            text_count = 0
            num_count = 0
            empty_count = 0
            for val in row:
                if val is None:
                    empty_count += 1
                elif isinstance(val, str):
                    # Check if it looks like a number
                    if val.strip().replace(".", "", 1).isdigit():
                        num_count += 1
                    else:
                        text_count += 1
                elif isinstance(val, (int, float)):
                    num_count += 1
                    
            score = text_count - num_count
            if text_count > 0 and score > best_score:
                best_score = score
                best_row = r_idx
                
        headers = {}
        # Fetch only the cells in the best_row up to max_col
        best_row_cells = list(ws.iter_rows(min_row=best_row, max_row=best_row, min_col=1, max_col=max_col, values_only=True))[0]
        for c_idx, val in enumerate(best_row_cells, start=1):
            if val is not None:
                headers[get_column_letter(c_idx)] = str(val)
        return headers

    def _infer_column_types(self, ws, headers: Dict[str, str]) -> Dict[str, str]:
        column_types = {}
        
        # Determine which row headers came from
        header_row = 1
        if headers:
            col_indices = [column_index_from_string(col_letter) for col_letter in headers.keys()]
            if col_indices:
                min_c = min(col_indices)
                max_c = max(col_indices)
                max_r = min(15, (ws.max_row or 1) + 1)
                
                rows_header_scan = list(ws.iter_rows(min_row=1, max_row=max_r, min_col=min_c, max_col=max_c))
                for r_idx, row in enumerate(rows_header_scan, start=1):
                    match_count = 0
                    for cell in row:
                        col_letter = get_column_letter(cell.column)
                        if col_letter in headers and str(cell.value or "") == headers[col_letter]:
                            match_count += 1
                    if match_count > 0:
                        header_row = r_idx
                        break
                    
        start_row = header_row + 1
        end_row = min((ws.max_row or 1), start_row + 50)
        
        if headers:
            col_indices = [column_index_from_string(col_letter) for col_letter in headers.keys()]
            min_c = min(col_indices)
            max_c = max(col_indices)
            
            # Read all cell values in a single iter_rows call
            rows_data = list(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_c, max_col=max_c))
            
            cells_by_col = {c: [] for c in col_indices}
            for row in rows_data:
                for cell in row:
                    if cell.column in cells_by_col:
                        cells_by_col[cell.column].append(cell)
                    
            for col_letter in headers.keys():
                c_idx = column_index_from_string(col_letter)
                types = []
                for cell in cells_by_col.get(c_idx, []):
                    val = cell.value
                    if val is None:
                        continue
                    
                    if isinstance(val, str) and val.startswith("="):
                        types.append("formula")
                    elif isinstance(val, (int, float)):
                        types.append("numeric")
                    elif isinstance(val, str):
                        if val.strip().replace(".", "", 1).isdigit():
                            types.append("numeric")
                        else:
                            # check if date
                            if len(val) >= 8 and ("-" in val or "/" in val):
                                types.append("date")
                            else:
                                types.append("text")
                    else:
                        import datetime
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            types.append("date")
                        else:
                            types.append("text")
                            
                if not types:
                    column_types[col_letter] = "text"
                else:
                    from collections import Counter
                    most_common = Counter(types).most_common(1)[0][0]
                    column_types[col_letter] = most_common
                    
        return column_types

    def _detect_locale(self, ws, active_cell: Optional[str] = None) -> str:
        # 1. Scan around the active cell first (e.g., 15x15 region) if provided
        if active_cell:
            try:
                col_letter, row_num = self._parse_cell_address(active_cell)
                min_row = max(1, row_num - 7)
                max_row = min(ws.max_row or 1, row_num + 7)
                min_col = max(1, col_letter - 7)
                max_col = min(ws.max_column or 1, col_letter + 7)
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=False):
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            in_quotes = False
                            for char in val:
                                if char == '"':
                                    in_quotes = not in_quotes
                                elif char == ';' and not in_quotes:
                                    return "eu"
                                elif char == ',' and not in_quotes:
                                    pass
            except Exception:
                pass

        # 2. Bounded scan of the first 100 rows and 20 columns as fallback
        max_r = min(ws.max_row or 1, 100)
        max_c = min(ws.max_column or 1, 20)
        for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c, values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    in_quotes = False
                    for char in val:
                        if char == '"':
                            in_quotes = not in_quotes
                        elif char == ';' and not in_quotes:
                            return "eu"
                        elif char == ',' and not in_quotes:
                            pass
        return "en"




class ExcelOperationValidator:
    """Validates formulas, checks for circular references, and corrects locale-specific elements."""

    def __init__(self):
        self.validator = ForgeValidator()

    def validate(self, op: dict, context: ExcelContext) -> ValidationResult:
        op_type = op.get("type", op.get("action", ""))

        if op_type in ("write_cell", "write_range", "excelmcp_fill_formula"):
            formula = op.get("formula", "")
            if formula:
                # 1. Circular reference check
                if context.active_cell.upper() in formula.upper():
                    return ValidationResult(
                        valid=False,
                        error=f"Circular reference detected: target cell {context.active_cell} referenced in formula '{formula}'",
                        op=op
                    )

                # Missing equals prefix correction
                if not formula.strip().startswith("="):
                    formula = "=" + formula.strip()
                    op["formula"] = formula

                # 2. Invoke Forge Formula Validator
                res = self.validator.validate_and_fix(formula, context.locale)
                
                # Apply corrected formula if available
                if res.get("corrected"):
                    op["formula"] = res["corrected"]

                if not res["valid"]:
                    if "corrected" in res:
                        # Re-run validation on corrected formula
                        sub_res = self.validator.validate_and_fix(res["corrected"], context.locale)
                        if sub_res["valid"]:
                            op["formula"] = sub_res["formula"]
                            return ValidationResult(valid=True, op=op)
                    return ValidationResult(
                        valid=False,
                        error=f"Formula validation error: {res['error']}",
                        op=op
                    )

                op["formula"] = res["formula"]

                # Recheck circular reference on corrected formula
                if context.active_cell.upper() in op["formula"].upper():
                    return ValidationResult(
                        valid=False,
                        error=f"Circular reference detected in formula: {context.active_cell} in '{op['formula']}'",
                        op=op
                    )

        return ValidationResult(valid=True, op=op)


class ExcelWriter:
    """Applies Excel operations preserving sheets, styles, number formats, and VBA macros."""

    def apply_operations(self, file_path: str, operations: List[dict]) -> dict:
        # STRATEGY: Try COM first if file is open in live Excel (windows only)
        try:
            import sys
            if sys.platform == "win32":
                from sidecar.parsers.excelmcp_bridge import excel_is_open
                has_cf = any(op.get("conditional_formatting") for op in operations)
                if excel_is_open(file_path) and not has_cf:
                    log.info("Excel is open and no conditional formatting requested. Applying operations via live COM...")
                    from sidecar.parsers.excelmcp_bridge import excelmcp_write_cell, excelmcp_write_range
                    applied = []
                    errors = []
                    for op in operations:
                        op_type = op.get("type", op.get("action", ""))
                        cell = op.get("cell")
                        formula = op.get("formula")
                        value = op.get("value")
                        
                        try:
                            if op_type == "write_cell":
                                res = excelmcp_write_cell(file_path, cell, value=value, formula=formula)
                                if res.get("ok"):
                                    applied.append(op)
                                else:
                                    errors.append(f"COM write_cell failed: {res.get('error')}")
                            elif op_type == "write_range":
                                range_spec = op.get("range") or cell
                                res = excelmcp_write_range(file_path, range_spec, values=op.get("values"), formulas=op.get("formulas"))
                                if res.get("ok"):
                                    applied.append(op)
                                else:
                                    errors.append(f"COM write_range failed: {res.get('error')}")
                        except Exception as e:
                            errors.append(f"COM operation error: {e}")
                    
                    if not errors:
                        return {"applied_count": len(applied), "errors": []}
        except Exception as e:
            log.warning(f"COM live Excel automation failed: {e}. Falling back to openpyxl formatting writer.")

        # Fallback: Write via openpyxl preserving formatting and macros
        try:
            wb = load_workbook(file_path, keep_vba=True)
            applied_count = 0
            errors = []
            
            for op in operations:
                try:
                    op_type = op.get("type", op.get("action", ""))
                    sheet_name = op.get("sheet", op.get("sheet_name"))
                    if sheet_name and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.active
                        
                    # Check sheet protection
                    if ws.protection.sheet or ws.protection.enabled:
                        formula_val = op.get("formula") or op.get("value")
                        raise PermissionError(
                            f"Sheet '{ws.title}' is protected. "
                            f"Please copy the formula manually from the Ghost Review Panel: {formula_val}"
                        )
                        
                    if op_type == "write_cell":
                        cell_ref = op.get("cell") or op.get("range")
                        target_cell = ws[cell_ref]
                        
                        existing_format = target_cell.number_format
                        
                        formula = op.get("formula")
                        if formula:
                            from sidecar.parsers.forge_bridge import validate_formula
                            val_res = validate_formula(formula, {"active_cell": cell_ref})
                            if not val_res["valid"]:
                                raise ValueError(f"Invalid Excel formula rejected: {formula}. Error: {val_res.get('error')}")
                            target_cell.value = val_res["corrected"]
                        elif op.get("value") is not None:
                            target_cell.value = op["value"]
                            
                        if op.get("bold"):
                            target_cell.font = Font(bold=True)
                            
                        # Preserve format if not explicitly changed
                        if not op.get("number_format"):
                            target_cell.number_format = existing_format
                        else:
                            target_cell.number_format = op["number_format"]
                            
                        applied_count += 1
                        
                    elif op_type == "write_range":
                        start_cell_ref = op.get("start_cell", op.get("cell", op.get("range", "A1")))
                        start_cell = ws[start_cell_ref]
                        start_row = start_cell.row
                        start_col = start_cell.column
                        
                        values = op.get("values", op.get("data", []))
                        for r_offset, row in enumerate(values):
                            for c_offset, value in enumerate(row):
                                if isinstance(value, str) and value.startswith("="):
                                    from sidecar.parsers.forge_bridge import validate_formula
                                    from openpyxl.utils import get_column_letter
                                    current_cell_ref = f"{get_column_letter(start_col+c_offset)}{start_row+r_offset}"
                                    val_res = validate_formula(value, {"active_cell": current_cell_ref})
                                    if not val_res["valid"]:
                                        raise ValueError(f"Invalid Excel formula in range rejected: {value}. Error: {val_res.get('error')}")
                                    value = val_res["corrected"]
                                ws.cell(row=start_row+r_offset, column=start_col+c_offset, value=value)
                                applied_count += 1
                                
                    elif op_type == "create_chart":
                        from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference
                        sheet_name, cell_range = _parse_range_spec(op.get("source_range"), ws.title)
                        ws_data = wb[sheet_name]
                        
                        # Parse range boundaries
                        parts = cell_range.upper().split(":")
                        min_cell = parts[0]
                        max_cell = parts[1] if len(parts) > 1 else parts[0]
                        min_col_str, min_row = parse_cell_ref(min_cell)
                        max_col_str, max_row = parse_cell_ref(max_cell)
                        
                        min_col = column_index_from_string(min_col_str)
                        max_col = column_index_from_string(max_col_str)
                        
                        ctype = op.get("chart_type", "column").lower()
                        if ctype in ("bar", "column"):
                            chart = BarChart()
                            chart.type = "col" if ctype == "column" else "bar"
                        elif ctype == "line":
                            chart = LineChart()
                        elif ctype == "pie":
                            chart = PieChart()
                        else:
                            chart = BarChart()
                            chart.type = "col"
                            
                        chart.title = op.get("title", "Chart")
                        data = Reference(ws_data, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                        chart.add_data(data, titles_from_data=True)
                        
                        t_sheet = op.get("target_sheet") or sheet_name
                        if t_sheet not in wb.sheetnames:
                            wb.create_sheet(t_sheet)
                        ws_target = wb[t_sheet]
                        
                        ws_target.add_chart(chart, f"{get_column_letter(max_col + 1)}{min_row}")
                        applied_count += 1
                        
                    elif op_type == "create_pivot":
                        sheet_name, cell_range = _parse_range_spec(op.get("source_range"), ws.title)
                        ws_src = wb[sheet_name]
                        
                        parts = cell_range.upper().split(":")
                        min_cell = parts[0]
                        _, min_row = parse_cell_ref(min_cell)
                        max_cell = parts[1] if len(parts) > 1 else parts[0]
                        max_col_str, _ = parse_cell_ref(max_cell)
                        max_col = column_index_from_string(max_col_str)
                        min_col_str, _ = parse_cell_ref(min_cell)
                        min_col = column_index_from_string(min_col_str)
                        
                        header_map = {}
                        for c in range(min_col, max_col + 1):
                            h = ws_src.cell(row=min_row, column=c).value
                            if h:
                                header_map[str(h)] = get_column_letter(c)
                                
                        t_sheet = op.get("target_sheet") or (sheet_name + "_Pivot")
                        if t_sheet in wb.sheetnames:
                            del wb[t_sheet]
                        ws_pivot = wb.create_sheet(t_sheet)
                        
                        pivot_headers = op.get("rows", []) + op.get("columns", []) + op.get("values", [])
                        for i, h in enumerate(pivot_headers, start=1):
                            ws_pivot.cell(row=1, column=i).value = h
                            
                        formula_row = 2
                        for val_col in op.get("values", []):
                            val_letter = header_map.get(val_col, "A")
                            for row_field in op.get("rows", []):
                                row_letter = header_map.get(row_field, "A")
                                ws_pivot.cell(row=formula_row, column=1).value = f"={row_field} totals"
                                formula = f"=SUMIF({sheet_name}!{row_letter}:{row_letter},\"<>\",{sheet_name}!{val_letter}:{val_letter})"
                                ws_pivot.cell(row=formula_row, column=len(pivot_headers)).value = formula
                                formula_row += 1
                        applied_count += 1
                        
                except Exception as e:
                    errors.append(str(e))
                    log.error(f"ExcelWriter operation failed: {e}")
                    
            if errors:
                return {"applied_count": applied_count, "errors": errors}
                
            # Atomic save
            tmp = file_path + ".kairo_tmp"
            wb.save(tmp)
            if os.path.exists(file_path):
                os.replace(tmp, file_path)
            else:
                os.rename(tmp, file_path)
                
            return {"applied_count": applied_count, "errors": []}
        except Exception as e:
            log.error(f"ExcelWriter failed: {traceback.format_exc()}")
            return {"applied_count": 0, "errors": [str(e)]}


def _parse_range_spec(range_spec: str, default_sheet: str) -> tuple[str, str]:
    if "!" in range_spec:
        sheet, cell_range = range_spec.split("1", 1) if "!" not in range_spec else range_spec.split("!", 1)
        return sheet.strip("'\""), cell_range
    return default_sheet, range_spec


def parse_cell_ref(ref: str) -> tuple[str, int]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", ref.strip())
    if m:
        return m.group(1).upper(), int(m.group(2))
    return "A", 1


# ---------------------------------------------------------------------------
# ExcelMaster — Unified Façade (required by DomainMasterRouter)
# Wraps ExcelContextExtractor + ExcelOperationValidator + ExcelWriter
# into the standard master interface: extract_context / build_prompt /
# validate_operations / apply_operations / get_schema_class.
# ---------------------------------------------------------------------------

class ExcelMaster:
    """
    Unified Excel Domain Master.
    Implements the standard interface used by DomainMasterRouter so every
    domain can be addressed uniformly.
    """

    def __init__(self):
        self._extractor = ExcelContextExtractor()
        self._validator = ExcelOperationValidator()
        self._writer = ExcelWriter()

    # --- Standard interface ---------------------------------------------------

    def extract_context(self, file_path: str, cursor_info=None) -> ExcelContext:
        """Extract spreadsheet context centered on the active cell."""
        active_cell = "A1"
        if cursor_info is not None:
            cell_str = str(cursor_info).strip()
            if cell_str:
                active_cell = cell_str
        return self._extractor.extract(file_path, active_cell)

    def build_prompt(
        self,
        user_prompt: str,
        doc_context: ExcelContext,
        mem_context: str,
        classification=None,
    ) -> str:
        """Build a fully-assembled Excel domain prompt with zero unreplaced variables."""
        import json

        active_cell = getattr(doc_context, "active_cell", "A1")
        active_sheet = getattr(doc_context, "active_sheet", "Sheet1")
        sheet_names = getattr(doc_context, "sheet_names", ["Sheet1"])
        cells = getattr(doc_context, "cells", [])
        headers = getattr(doc_context, "headers", {})
        named_ranges = getattr(doc_context, "named_ranges", {})
        locale = getattr(doc_context, "locale", "en")
        file_path = getattr(doc_context, "file_path", "Unknown")

        column_types_dict: dict = {}
        for c in cells:
            ref = c.get("address", "A1")
            col_letter = "".join(ch for ch in ref if ch.isalpha())
            val = c.get("value")
            if val is not None and col_letter and col_letter not in column_types_dict:
                try:
                    float(val)
                    column_types_dict[col_letter] = "Number"
                except (ValueError, TypeError):
                    column_types_dict[col_letter] = "Text"

        headers_list = list(headers.values()) if isinstance(headers, dict) else []

        mem_str = mem_context or "No writing preferences learned yet. Use professional defaults."

        return f"""SYSTEM:
You are the Excel Spreadsheet Master for Kairo Phantom. You have deep expertise in Excel formulas,
financial modeling, data analysis, and spreadsheet best practices.

CONSTRAINTS:
- Output ONLY valid JSON matching ExcelResponse schema. No prose. No explanation. No code fences.
- NEVER modify cells not explicitly targeted
- NEVER auto-calculate or evaluate formulas — write them exactly as strings starting with =
- Inside JSON string values, NEVER use unescaped double quotes — use single quotes for text args.

FORMULA RULES:
1. ALL formulas MUST start with =
2. Match locale: en=commas =SUM(A1,A2), eu=semicolons =SUM(A1;A2)
3. Always include the 4th argument in VLOOKUP: =VLOOKUP(A1,B:C,2,FALSE)
4. Use absolute references for lookup ranges: =VLOOKUP(A1,$B$1:$C$100,2,FALSE)
5. Prefer XLOOKUP over VLOOKUP for Excel 365

=== APP CONTEXT ===
Application Name: Microsoft Excel
Application Type: Spreadsheet
File Path: {file_path}

=== DOCUMENT CONTEXT ===
Active cell: {active_cell} on sheet: {active_sheet}
Available sheets: {json.dumps(list(sheet_names) if sheet_names else ["Sheet1"])}
Detected headers: {json.dumps(headers_list)}
Named ranges: {json.dumps(named_ranges)}
Column types: {json.dumps(column_types_dict)}
Locale: {locale} (en=comma separator, eu=semicolon separator)

=== MEMORY CONTEXT ===
User Writing Preferences:
{mem_str}

=== INTENT CLASSIFICATION ===
Intent: {getattr(classification, "task_type", "insert") if classification else "insert"}

REMINDER: Your entire response must be a single JSON object. First character must be {{. Last character must be }}.
USER INSTRUCTION:
{user_prompt}
OUTPUT (JSON only):
"""

    def validate_operations(self, raw_response, doc_context: ExcelContext) -> list:
        """Validate Excel operations including circular reference and locale checks."""
        validated = []
        ops = getattr(raw_response, "operations", [])
        if isinstance(ops, list):
            for op in ops:
                op_dict = op.model_dump() if hasattr(op, "model_dump") else dict(op)
                result = self._validator.validate(op_dict, doc_context)
                if result.valid:
                    validated.append(result.op)
                else:
                    log.warning(f"ExcelMaster rejected op: {result.error}")
        return validated

    def apply_operations(self, file_path: str, operations: list) -> dict:
        """Write validated operations to the .xlsx file atomically."""
        return self._writer.apply_operations(file_path, operations)

    def get_schema_class(self):
        """Return the Pydantic schema class for LLM structured output."""
        from sidecar.schemas.xlsx_schema import ExcelResponse
        return ExcelResponse
