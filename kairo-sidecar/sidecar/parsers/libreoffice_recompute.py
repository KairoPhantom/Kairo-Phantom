# ==============================================================================
# DISPOSITION: REAL (No Mock Patterns)
# ------------------------------------------------------------------------------
# This module uses LibreOffice (soffice) headless mode to force-recalculate
# an .xlsx file's formulas and return the computed cell values. It does NOT
# mock, stub, or simulate LibreOffice. If soffice is unavailable, it raises
# RuntimeError with a clear message.
# ==============================================================================

"""
LibreOffice Recompute — Forces formula recalculation via headless soffice.

Uses subprocess to call:
    soffice --headless --nologo --nofirststartwizard --convert-to xlsx --outdir <tmp> <input.xlsx>

A custom user profile with registrymodifications.xcu is created to force
LibreOffice to recalculate all formulas on load (OOXMLRecalcMode=1,
ODFRecalcMode=1). Without this, LibreOffice may skip recalculation and
return cached (stale) values.

Then parses the recomputed .xlsx with openpyxl (data_only=True) to extract
calculated cell values.

Returns: Dict[cell_ref, computed_value]
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any

log = logging.getLogger("kairo-sidecar.libreoffice_recompute")

# Cache the soffice path so we don't re-check every call
_soffice_path: str | None = None
_soffice_checked: bool = False

# Registry XML that forces LibreOffice Calc to recalculate formulas on load.
# Without this, soffice --convert-to may return stale cached values.
_REGISTRY_MODIFICATIONS = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>1</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>1</value></prop></item>
</oor:items>
"""


def _find_soffice() -> str:
    """Locate the soffice binary. Returns the full path or raises RuntimeError."""
    global _soffice_path, _soffice_checked
    if _soffice_checked:
        if _soffice_path is None:
            raise RuntimeError(
                "LibreOffice (soffice) is not available on this system. "
                "Install it with: apt-get install libreoffice-calc  (or equivalent)."
            )
        return _soffice_path

    _soffice_checked = True

    # Check common locations
    candidates = [
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/opt/libreoffice/program/soffice",
        "/snap/bin/libreoffice",
    ]
    for c in candidates:
        if os.path.isfile(c) and os.access(c, os.X_OK):
            _soffice_path = c
            log.info(f"Found soffice at: {c}")
            return _soffice_path

    # Try PATH lookup
    found = shutil.which("soffice")
    if found:
        _soffice_path = found
        log.info(f"Found soffice via PATH: {found}")
        return _soffice_path

    _soffice_path = None
    raise RuntimeError(
        "LibreOffice (soffice) is not available on this system. "
        "Install it with: apt-get install libreoffice-calc  (or equivalent)."
    )


def _create_profile_with_recalc(profile_dir: str) -> str:
    """
    Create a LibreOffice user profile directory with registrymodifications.xcu
    that forces formula recalculation on load.

    Returns the profile directory path.
    """
    user_dir = os.path.join(profile_dir, "user")
    os.makedirs(user_dir, exist_ok=True)
    registry_path = os.path.join(user_dir, "registrymodifications.xcu")
    with open(registry_path, "w") as f:
        f.write(_REGISTRY_MODIFICATIONS)
    return profile_dir


def recompute_xlsx(
    input_path: str,
    timeout: int = 60,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """
    Force-recalculate an .xlsx file using headless LibreOffice and return
    computed cell values.

    Args:
        input_path: Path to the input .xlsx file.
        timeout: Subprocess timeout in seconds (default 60).
        sheet_name: If specified, only return cells from this sheet.
                     If None, returns cells from all sheets (prefixed with sheet name).

    Returns:
        Dict mapping cell references to computed values.
        When sheet_name is None, keys are "SheetName!A1" format.
        When sheet_name is specified, keys are "A1" format.

    Raises:
        RuntimeError: If soffice is not available or the conversion fails.
        FileNotFoundError: If the input file does not exist.
    """
    soffice = _find_soffice()

    input_file = Path(input_path)
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    # Create a temp directory for the output and profile
    with tempfile.TemporaryDirectory(prefix="kairo_lo_recompute_") as tmpdir:
        # Create a user profile with forced recalculation
        profile_dir = os.path.join(tmpdir, "lo_profile")
        _create_profile_with_recalc(profile_dir)

        # Use a separate output directory to avoid overwriting the input
        outdir = os.path.join(tmpdir, "output")
        os.makedirs(outdir, exist_ok=True)

        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--norestore",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "xlsx",
            "--outdir",
            outdir,
            str(input_file),
        ]

        log.info(f"Running soffice recompute: {' '.join(cmd)}")
        start_time = time.monotonic()

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        except subprocess.TimeoutExpired:
            raise RuntimeError(
                f"LibreOffice subprocess timed out after {timeout}s. "
                "The file may be too complex or soffice may be stuck."
            )

        elapsed = time.monotonic() - start_time
        log.info(f"soffice completed in {elapsed:.2f}s, exit_code={result.returncode}")

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed (exit {result.returncode}). "
                f"stderr: {result.stderr.strip()[:500]}"
            )

        # The output file has the same basename as the input
        output_file = Path(outdir) / input_file.name

        if not output_file.exists():
            # Sometimes soffice creates the file with a different case or extension
            xlsx_files = list(Path(outdir).glob("*.xlsx"))
            if xlsx_files:
                output_file = xlsx_files[0]
            else:
                raise RuntimeError(
                    f"LibreOffice did not produce an output .xlsx file. "
                    f"stdout: {result.stdout.strip()[:300]}, "
                    f"stderr: {result.stderr.strip()[:300]}"
                )

        # Parse the recomputed file with openpyxl in data_only mode
        # to get the cached (computed) values
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required but not installed.")

        wb = openpyxl.load_workbook(str(output_file), data_only=True)
        computed: dict[str, Any] = {}

        sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
        for sname in sheets_to_process:
            if sname not in wb.sheetnames:
                raise RuntimeError(
                    f"Sheet '{sname}' not found in recomputed workbook. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sname]
            prefix = "" if sheet_name else f"{sname}!"
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        computed[f"{prefix}{cell.coordinate}"] = cell.value

        wb.close()
        log.info(f"Recomputed {len(computed)} cells in {elapsed:.2f}s")
        return computed


def recompute_xlsx_with_timing(
    input_path: str,
    timeout: int = 60,
    sheet_name: str | None = None,
) -> tuple[dict[str, Any], float]:
    """
    Same as recompute_xlsx but also returns the elapsed time in seconds.

    Returns:
        Tuple of (computed_values_dict, elapsed_seconds).
    """
    start = time.monotonic()
    result = recompute_xlsx(input_path, timeout=timeout, sheet_name=sheet_name)
    elapsed = time.monotonic() - start
    return result, elapsed


def soffice_available() -> bool:
    """Check if soffice is available without raising."""
    try:
        _find_soffice()
        return True
    except RuntimeError:
        return False