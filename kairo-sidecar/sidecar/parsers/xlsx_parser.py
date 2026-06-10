import logging
from pathlib import Path
from typing import Optional

log = logging.getLogger("kairo-sidecar.xlsx_parser")

def parse_xlsx(file_path: str, active_cell: Optional[str] = None) -> dict:
    """
    Read Excel file and return surrounding context grid and named ranges.
    """
    try:
        import openpyxl
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {path}"}

    active_sheet = None
    try:
        import win32com.client
        import pythoncom
        import subprocess
        pythoncom.CoInitialize()
        
        excel_running = False
        try:
            out = subprocess.run(["tasklist", "/FI", "IMAGENAME eq excel.exe"], capture_output=True, text=True)
            excel_running = "excel.exe" in out.stdout.lower()
        except Exception:
            pass
            
        if excel_running:
            target_path = str(path.resolve())
            try:
                wb_com = win32com.client.GetObject(target_path)
                xl = wb_com.Application
                active_sheet = wb_com.ActiveSheet.Name
                active_cell = xl.ActiveCell.Address.replace("$", "")
                log.info(f"parse_xlsx COM detected active sheet via moniker: {active_sheet}, cell: {active_cell}")
            except Exception:
                try:
                    xl = win32com.client.GetActiveObject("Excel.Application")
                    for wb in xl.Workbooks:
                        if str(Path(wb.FullName).resolve()) == target_path:
                            active_sheet = wb.ActiveSheet.Name
                            active_cell = xl.ActiveCell.Address.replace("$", "")
                            log.info(f"parse_xlsx COM detected active sheet via fallback: {active_sheet}, cell: {active_cell}")
                            break
                except Exception:
                    pass
    except Exception as e:
        log.debug(f"parse_xlsx could not get Excel active info via COM: {e}")

    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
        if active_sheet and active_sheet in wb.sheetnames:
            ws = wb[active_sheet]
        else:
            ws = wb.active

        # Parse active cell (e.g. "G5")
        if active_cell:
            col_str = ''.join(c for c in active_cell if c.isalpha())
            row_digits = ''.join(c for c in active_cell if c.isdigit())
            row = int(row_digits) if row_digits else 1
            col = column_index_from_string(col_str) if col_str else 1
        else:
            row, col = 1, 1

        # Extract 10x10 window around active cell
        r_start = max(1, row - 5)
        r_end = min(ws.max_row, row + 5)
        c_start = max(1, col - 5)
        c_end = min(ws.max_column, col + 5)

        grid = []
        for r in range(r_start, r_end + 1):
            row_data = []
            for c in range(c_start, c_end + 1):
                cell = ws.cell(row=r, column=c)
                row_data.append({
                    "ref": f"{get_column_letter(c)}{r}",
                    "value": str(cell.value) if cell.value is not None else "",
                    "formula": str(cell.value) if str(cell.value).startswith("=") else "",
                    "is_active": (r == row and c == col),
                })
            grid.append(row_data)

        # Column headers (row 1)
        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                headers[get_column_letter(c)] = str(val)

        # Named ranges
        try:
            named_ranges = list(wb.defined_names.keys())
        except Exception:
            named_ranges = []

        return {
            "active_cell": active_cell or "A1",
            "active_row": row,
            "active_col": col,
            "sheet_name": ws.title,
            "sheet_names": wb.sheetnames,
            "grid": grid,
            "headers": headers,
            "named_ranges": named_ranges,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }
    except Exception as e:
        log.error(f"Failed to parse Excel file {file_path}: {e}")
        return {"error": f"Failed to parse Excel: {e}"}


def get_workbook_blueprint(file_path: str) -> dict:
    """
    Get structural overview of an Excel workbook.
    """
    from sidecar.parsers.excelmcp_bridge import get_workbook_blueprint as get_bp
    res = get_bp(file_path)
    if res.get("ok"):
        return res["data"]
    return {"sheets": [], "named_ranges": [], "total_sheets": 0, "total_named_ranges": 0}


def write_xlsx_with_formatting(file_path: str, operations: list[dict]) -> dict:
    """
    Write Excel operations WITH formatting preservation.
    Each operation: {cell, formula?, value?, number_format?, bold?, conditional_formatting?}
    conditional_formatting: {type: "cell_is"|"data_bar"|"color_scale", operator?, threshold?, fill_color?}
    """
    from sidecar.parsers.excelmcp_bridge import excel_com_lease, save_workbook_safely
    log.info(f"write_xlsx_with_formatting: file_path={file_path}, operations={operations}")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        active_sheet = None
        try:
            import win32com.client
            import pythoncom
            import subprocess
            pythoncom.CoInitialize()
            
            excel_running = False
            try:
                out = subprocess.run(["tasklist", "/FI", "IMAGENAME eq excel.exe"], capture_output=True, text=True)
                excel_running = "excel.exe" in out.stdout.lower()
            except Exception:
                pass
                
            if excel_running:
                target_path = str(path.resolve())
                try:
                    wb_com = win32com.client.GetObject(target_path)
                    active_sheet = wb_com.ActiveSheet.Name
                    log.info(f"write_xlsx_with_formatting COM detected active sheet: {active_sheet}")
                except Exception:
                    try:
                        xl = win32com.client.GetActiveObject("Excel.Application")
                        for wb_check in xl.Workbooks:
                            if str(Path(wb_check.FullName).resolve()) == target_path:
                                active_sheet = wb_check.ActiveSheet.Name
                                log.info(f"write_xlsx_with_formatting fallback COM active sheet: {active_sheet}")
                                break
                    except Exception:
                        pass
        except Exception as e:
            log.debug(f"write_xlsx_with_formatting COM active sheet detection failed: {e}")

        with excel_com_lease(file_path):
            wb = openpyxl.load_workbook(str(path), data_only=False)
            if active_sheet and active_sheet in wb.sheetnames:
                ws = wb[active_sheet]
            else:
                ws = wb.active

            written = 0
            for op in operations:
                cell_ref = op.get("cell")
                if not cell_ref:
                    continue

                formula = op.get("formula") or ""
                value = op.get("value")
                num_fmt = op.get("number_format")
                is_bold = op.get("bold")
                cf_config = op.get("conditional_formatting")

                # ── Conditional Formatting path (structured) ───────────────────
                if cf_config and isinstance(cf_config, dict):
                    cf_type = cf_config.get("type", "")
                    try:
                        if cf_type == "cell_is":
                            from openpyxl.formatting.rule import CellIsRule
                            operator = cf_config.get("operator", "greaterThan")
                            threshold = cf_config.get("threshold", 0)
                            fill_color = cf_config.get("fill_color", "C6EFCE")
                            fill = PatternFill(
                                start_color=fill_color, end_color=fill_color,
                                fill_type="solid"
                            )
                            rule = CellIsRule(
                                operator=operator,
                                formula=[str(threshold)],
                                fill=fill
                            )
                            ws.conditional_formatting.add(cell_ref, rule)
                            written += 1
                            continue

                        elif cf_type == "data_bar":
                            try:
                                from openpyxl.formatting.rule import DataBarRule
                                rule = DataBarRule(
                                    start_type="min", start_value=0,
                                    end_type="max", end_value=100,
                                    color="638EC6"
                                )
                            except (ImportError, TypeError):
                                # Older openpyxl — use ColorScaleRule as fallback
                                from openpyxl.formatting.rule import ColorScaleRule
                                rule = ColorScaleRule(
                                    start_type="min", start_color="FFFFFF",
                                    end_type="max", end_color="638EC6"
                                )
                            ws.conditional_formatting.add(cell_ref, rule)
                            written += 1
                            continue

                        elif cf_type == "color_scale":
                            from openpyxl.formatting.rule import ColorScaleRule
                            rule = ColorScaleRule(
                                start_type="min", start_color="F8696B",
                                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                end_type="max", end_color="63BE7B"
                            )
                            ws.conditional_formatting.add(cell_ref, rule)
                            written += 1
                            continue
                    except Exception as e:
                        log.warning(f"Structured conditional formatting failed ({cf_type}): {e}")

                # ── Range of cells ─────────────────────────────────────────────
                if ":" in cell_ref:
                    # Legacy heuristic: formula containing >, <, = → treat as CF
                    if formula and any(ch in formula for ch in [">", "<"]):
                        try:
                            from openpyxl.formatting.rule import CellIsRule
                            fill_color = "FFC7CE"  # default red
                            if ">" in formula:
                                fill_color = "C6EFCE"  # green for greater-than
                            fill = PatternFill(
                                start_color=fill_color, end_color=fill_color,
                                fill_type="solid"
                            )
                            operator = "greaterThan" if ">" in formula else "lessThan"
                            formula_val = formula.split(">")[-1] if ">" in formula else formula.split("<")[-1]
                            rule = CellIsRule(operator=operator, formula=[formula_val.strip()], fill=fill)
                            ws.conditional_formatting.add(cell_ref, rule)
                            written += 1
                            continue
                        except Exception as e:
                            log.warning(f"Legacy CF heuristic failed: {e}")

                    # Normal range value/formula write
                    # ws[range] may return a flat tuple of Cells (single row/col)
                    # or a tuple of row-tuples. Normalise to 2-D rows×cols.
                    raw = ws[cell_ref]
                    from openpyxl.cell.cell import Cell as _Cell
                    if raw and isinstance(raw[0], _Cell):
                        rows_iter = (raw,)  # single-row flat tuple → wrap
                    else:
                        rows_iter = raw
                    for row in rows_iter:
                        for c in row:
                            if formula:
                                c.value = formula if formula.startswith("=") else f"={formula}"
                            elif value is not None and value != "":
                                c.value = value
                            if num_fmt:
                                c.number_format = num_fmt
                            if is_bold:
                                existing = c.font
                                c.font = Font(
                                    bold=True,
                                    name=existing.name if existing else "Calibri",
                                    size=existing.size if existing else 11,
                                )
                            written += 1

                # ── Single cell ────────────────────────────────────────────────
                else:
                    cell = ws[cell_ref]
                    if formula:
                        cell.value = formula if formula.startswith("=") else f"={formula}"
                    elif value is not None and value != "":
                        cell.value = value

                    if num_fmt:
                        cell.number_format = num_fmt

                    if is_bold:
                        existing = cell.font
                        cell.font = Font(
                            bold=True,
                            name=existing.name if existing else "Calibri",
                            size=existing.size if existing else 11,
                        )
                    written += 1

            save_workbook_safely(wb, path)
        return {"ok": True, "data": {"cells_written": written}}
    except Exception as e:
        log.error(f"Failed write_xlsx_with_formatting: {e}")
        return {"error": str(e)}

