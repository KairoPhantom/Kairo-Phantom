"""
ExcelMcp Bridge — Live Excel COM automation for Kairo Phantom Domain 2.
======================================================================
Provides a 3-tier injection ladder:
  1. excel-mcp-server CLI  — live COM injection (best quality, requires server)
  2. win32com direct        — live COM without server (Windows + Excel open)
  3. openpyxl file-based   — always available, formatting-preserving

All public functions return dicts matching Kairo's sidecar JSON envelope:
  {"ok": bool, "data": {...}, "error": str | None}
"""

from __future__ import annotations

import json
import logging
import re
import shutil
import subprocess
import tempfile
import time
import traceback
from contextlib import contextmanager
from pathlib import Path
from typing import Any

log = logging.getLogger("kairo-sidecar.excelmcp_bridge")

# ──────────────────────────────────────────────────────────────────────────────
# Capability detection
# ──────────────────────────────────────────────────────────────────────────────


def excelmcp_available() -> bool:
    """True if the excel-mcp-server CLI is in PATH."""
    return shutil.which("excel-mcp-server") is not None


def win32com_available() -> bool:
    """True if win32com (pywin32) is importable."""
    try:
        import win32com.client  # noqa: F401

        return True
    except ImportError:
        return False


def excel_is_open(file_path: str) -> bool:
    """
    Windows-only: check if Excel has the given file open via COM.
    Returns False on non-Windows or any error — always fail-safe.
    """
    try:
        import sys

        if sys.platform != "win32":
            return False
        if not win32com_available():
            return False

        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()

        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return False

        target = Path(file_path).resolve()
        try:
            for wb in xl.Workbooks:
                if Path(wb.FullName).resolve() == target:
                    return True
        except Exception:
            pass
        return False
    except Exception:
        return False


@contextmanager
def excel_com_lease(file_path: str):
    """
    Context manager to release Excel's file lock.
    If the file is open in Excel, it saves and closes it via COM,
    yields to perform openpyxl operations, and reopens it in Excel.
    """
    import sys

    if sys.platform != "win32" or not win32com_available():
        yield
        return

    import win32com.client
    import pythoncom
    import subprocess

    pythoncom.CoInitialize()
    xl = None
    target_wb = None
    was_open = False
    target = Path(file_path).resolve()

    try:
        excel_running = False
        try:
            out = subprocess.run(
                ["tasklist", "/FI", "IMAGENAME eq excel.exe"], capture_output=True, text=True
            )
            excel_running = "excel.exe" in out.stdout.lower()
        except Exception:
            pass

        if excel_running:
            try:
                target_wb = win32com.client.GetObject(str(target))
                xl = target_wb.Application
                was_open = True
            except Exception:
                try:
                    xl = win32com.client.GetActiveObject("Excel.Application")
                    for wb in xl.Workbooks:
                        if Path(wb.FullName).resolve() == target:
                            target_wb = wb
                            was_open = True
                            break
                except Exception:
                    pass
    except Exception:
        pass

    if was_open and target_wb is not None and xl is not None:
        try:
            log.info("excel_com_lease: Saving and closing workbook %s via COM...", target.name)
            target_wb.Save()
            target_wb.Close(SaveChanges=True)
            time.sleep(1.5)
        except Exception as e:
            log.warning("excel_com_lease: Failed to close workbook via COM: %s", e)

    try:
        yield
    finally:
        if was_open:
            try:
                # Check if the Excel application process is still alive
                try:
                    xl.Visible = True
                except Exception:
                    # Excel was killed or COM reference is dead. Launch a new instance!
                    log.info(
                        "excel_com_lease: Excel application reference is dead. Launching new Excel instance..."
                    )
                    xl = win32com.client.Dispatch("Excel.Application")

                log.info("excel_com_lease: Reopening workbook %s in Excel...", target.name)
                for attempt in range(5):
                    try:
                        xl.Workbooks.Open(str(target))
                        xl.Visible = True
                        time.sleep(1.5)
                        break
                    except Exception as e:
                        if attempt == 4:
                            raise e
                        log.warning(
                            "excel_com_lease: Reopen attempt %d failed: %s. Retrying in 1.5s...",
                            attempt + 1,
                            e,
                        )
                        time.sleep(1.5)
            except Exception as e:
                log.warning("excel_com_lease: Failed to reopen workbook: %s", e)


def save_workbook_safely(wb, path):
    """
    Saves an openpyxl Workbook to path. If a PermissionError is encountered,
    it force-kills Excel to release the file lock and retries the save.
    """
    try:
        wb.save(str(path))
    except PermissionError as pe:
        log.warning(
            "save_workbook_safely: PermissionError encountered: %s. Force-killing Excel to release lock and retrying...",
            pe,
        )
        import os

        os.system("taskkill /F /IM EXCEL.EXE /T >nul 2>&1")
        time.sleep(2)
        wb.save(str(path))


# ──────────────────────────────────────────────────────────────────────────────
# Column helpers (Pi-style)
# ──────────────────────────────────────────────────────────────────────────────


def col_to_idx(col: str) -> int:
    """A→0, B→1, …, Z→25, AA→26, …"""
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def idx_to_col(idx: int) -> str:
    """0→A, 1→B, …, 25→Z, 26→AA, …"""
    result = ""
    n = idx
    while n >= 0:
        result = chr(ord("A") + n % 26) + result
        n = n // 26 - 1
    return result


def parse_cell_ref(ref: str) -> tuple[str, int]:
    """Parse 'C7' → ('C', 7). Returns ('A', 1) on failure."""
    m = re.match(r"^([A-Za-z]+)(\d+)$", ref.strip())
    if m:
        return m.group(1).upper(), int(m.group(2))
    return "A", 1


# ──────────────────────────────────────────────────────────────────────────────
# Workbook blueprint
# ──────────────────────────────────────────────────────────────────────────────


def get_workbook_blueprint(file_path: str) -> dict:
    """
    Get structural overview of an Excel workbook via openpyxl.
    Returns:
        {ok: True, data: {sheets: [...], named_ranges: [...], total_sheets: int, ...}}
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(
                {
                    "name": name,
                    "max_row": ws.max_row or 0,
                    "max_col": ws.max_column or 0,
                    "has_tables": bool(getattr(ws, "tables", {})),
                }
            )
        try:
            named_ranges = list(wb.defined_names.keys())
        except Exception:
            named_ranges = []

        tables: list[dict] = []
        # Scan each sheet's tables (requires read_only=False for some openpyxl versions)
        try:
            wb2 = openpyxl.load_workbook(str(path), data_only=True)
            for name in wb2.sheetnames:
                ws2 = wb2[name]
                for tname, tobj in (getattr(ws2, "tables", {}) or {}).items():
                    tables.append({"name": tname, "sheet": name, "range": getattr(tobj, "ref", "")})
        except Exception:
            pass

        return {
            "ok": True,
            "data": {
                "sheets": sheets,
                "named_ranges": named_ranges,
                "tables": tables,
                "total_sheets": len(sheets),
                "total_named_ranges": len(named_ranges),
            },
        }
    except ImportError:
        return _error("openpyxl not installed — run: pip install openpyxl")
    except Exception:
        return _error(traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────────
# Read operations
# ──────────────────────────────────────────────────────────────────────────────


def excelmcp_read_range(file_path: str, range_spec: str, mode: str = "compact") -> dict:
    """
    Read a cell range from an xlsx file.
    range_spec: "A1:D10" or "Sheet1!A1:D10"
    mode: "compact" (markdown), "csv", or "detailed"
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=False)

        sheet_name, cell_range = _parse_range_spec(range_spec, wb.active.title)
        if sheet_name not in wb.sheetnames:
            return _error(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]

        cells = []
        # ws[range] may return a flat tuple of Cells for single-row/col ranges.
        # Normalise to 2-D rows×cols so we can always do: for row in ...: for cell in row:
        raw_range = ws[cell_range]
        from openpyxl.cell.cell import Cell as _Cell

        rows_2d = (raw_range,) if (raw_range and isinstance(raw_range[0], _Cell)) else raw_range
        for row in rows_2d:
            row_data = []
            for cell in row:
                val = cell.value
                formula = ""
                value_str = ""
                if isinstance(val, str) and val.startswith("="):
                    formula = val
                elif val is not None:
                    value_str = str(val)
                row_data.append(
                    {
                        "ref": cell.coordinate,
                        "value": value_str,
                        "formula": formula,
                        "number_format": cell.number_format or "General",
                    }
                )
            cells.append(row_data)

        if mode == "compact":
            # Markdown table
            lines = []
            for row in cells:
                lines.append(" | ".join(c["formula"] or c["value"] or "" for c in row))
            data_str = "\n".join(lines)
            return {"ok": True, "data": {"mode": "compact", "content": data_str, "cells": cells}}
        elif mode == "csv":
            lines = []
            for row in cells:
                lines.append(",".join(f'"{c["formula"] or c["value"]}"' for c in row))
            return {
                "ok": True,
                "data": {"mode": "csv", "content": "\n".join(lines), "cells": cells},
            }
        else:
            return {"ok": True, "data": {"mode": "detailed", "cells": cells}}
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────────
# Write operations
# ──────────────────────────────────────────────────────────────────────────────


def excelmcp_write_cell(
    file_path: str,
    cell: str,
    value: Any = None,
    formula: str | None = None,
    number_format: str | None = None,
    bold: bool = False,
) -> dict:
    """
    Write a value or formula to a single cell, preserving all other formatting.
    Tries: win32com live COM → openpyxl file-based.
    """
    # Strategy 1: Live COM (if Excel has file open)
    if excel_is_open(file_path):
        result = _win32com_write_cell(file_path, cell, value, formula)
        if result["ok"]:
            return result
        log.warning("win32com write failed, falling back to openpyxl: %s", result.get("error"))

    # Strategy 2: openpyxl file-based
    return _openpyxl_write_cell(file_path, cell, value, formula, number_format, bold)


def excelmcp_write_range(
    file_path: str,
    range_spec: str,
    values: list[list] | None = None,
    formulas: list[list[str]] | None = None,
) -> dict:
    """
    Write values/formulas to a range with formatting preservation.
    values and formulas are row-major 2D arrays. If formulas provided, they take priority.
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")
    try:
        with excel_com_lease(file_path):
            import openpyxl

            wb = openpyxl.load_workbook(str(path), data_only=False)
            sheet_name, cell_range = _parse_range_spec(range_spec, wb.active.title)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            rows_written = 0
            for row_idx, row in enumerate(ws[cell_range]):
                for col_idx, cell in enumerate(row):
                    if formulas and row_idx < len(formulas) and col_idx < len(formulas[row_idx]):
                        cell.value = formulas[row_idx][col_idx]
                    elif values and row_idx < len(values) and col_idx < len(values[row_idx]):
                        cell.value = values[row_idx][col_idx]
                rows_written += 1

            save_workbook_safely(wb, path)
            return {"ok": True, "data": {"rows_written": rows_written, "range": range_spec}}
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


def excelmcp_fill_formula(file_path: str, formula: str, fill_range: str) -> dict:
    """
    AutoFill a formula across a range — relative cell references are adjusted
    per row/col offset from the formula's starting position.
    E.g. formula='=A2-B2', fill_range='C2:C101' → writes =A2-B2, =A3-B3, ...
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")
    try:
        with excel_com_lease(file_path):
            import openpyxl
            from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: F401

            wb = openpyxl.load_workbook(str(path), data_only=False)
            sheet_name, cell_range = _parse_range_spec(fill_range, wb.active.title)
            ws = wb[sheet_name]

            written = 0
            # Determine starting row/col from cell_range start
            start_cell = cell_range.split(":")[0] if ":" in cell_range else cell_range
            start_col_str, start_row = parse_cell_ref(start_cell)
            start_col_idx = column_index_from_string(start_col_str)

            for row in ws[cell_range]:
                for cell in row:
                    row_offset = cell.row - start_row
                    col_offset = column_index_from_string(cell.column_letter) - start_col_idx
                    adjusted = _adjust_formula(formula, row_offset, col_offset)
                    cell.value = adjusted
                    written += 1

            save_workbook_safely(wb, path)
            return {"ok": True, "data": {"cells_written": written, "range": fill_range}}
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────────
# Creation operations
# ──────────────────────────────────────────────────────────────────────────────


def excelmcp_create_chart(
    file_path: str,
    source_range: str,
    chart_type: str,
    title: str,
    target_sheet: str | None = None,
) -> dict:
    """
    Create an Excel chart from a data range using openpyxl.
    chart_type: "bar", "column", "line", "pie", "scatter", "area"
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")

    # Direct COM chart creation (if Excel has file open)
    if excel_is_open(file_path):
        try:
            import win32com.client
            import pythoncom

            pythoncom.CoInitialize()

            target = Path(file_path).resolve()
            xl = None
            wb = None
            try:
                wb = win32com.client.GetObject(str(target))
                xl = wb.Application
            except Exception:
                xl = win32com.client.GetActiveObject("Excel.Application")
                for w in xl.Workbooks:
                    if Path(w.FullName).resolve() == target:
                        wb = w
                        break

            if wb is not None:
                sheet_name, cell_range = _parse_range_spec(source_range, wb.ActiveSheet.Name)
                ws_data = wb.Sheets(sheet_name)

                # Target sheet
                tsheet = target_sheet if target_sheet else sheet_name
                ws_target = wb.Sheets(tsheet)

                # Add chart object
                parts = cell_range.upper().split(":")
                max_cell = parts[1] if len(parts) > 1 else parts[0]
                max_col_str, min_row = parse_cell_ref(max_cell)
                max_col = col_to_idx(max_col_str) + 1

                anchor_col = idx_to_col(max_col)  # one column right of data

                # Line chart is type 4 (xlLine), Column chart is type 51 (xlColumnClustered)
                xl_chart_type = 4 if chart_type.lower() == "line" else 51
                chart_shape = ws_target.Shapes.AddChart2(Style=-1, XlChartType=xl_chart_type)
                chart = chart_shape.Chart

                # Set source data
                src_range = ws_data.Range(cell_range)
                chart.SetSourceData(Source=src_range)

                chart.HasTitle = True
                chart.ChartTitle.Text = title

                # Position chart near anchor
                anchor_cell = ws_target.Range(f"{anchor_col}{min_row}")
                chart_shape.Left = anchor_cell.Left
                chart_shape.Top = anchor_cell.Top

                wb.Save()
                log.info("Direct win32com chart creation OK: %s '%s'", chart_type, title)
                return {
                    "ok": True,
                    "data": {
                        "chart_type": chart_type,
                        "title": title,
                        "positioned_at": f"{anchor_col}{min_row}",
                        "sheet": tsheet,
                        "backend": "win32com",
                    },
                }
        except Exception as e:
            log.warning("Direct COM chart creation failed: %s. Falling back to openpyxl...", e)

    try:
        with excel_com_lease(file_path):
            import openpyxl
            from openpyxl.chart import (
                BarChart,
                LineChart,
                PieChart,
                ScatterChart,
                AreaChart,
                Reference,
            )

            wb = openpyxl.load_workbook(str(path))
            sheet_name, cell_range = _parse_range_spec(source_range, wb.active.title)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.active.title
            ws_data = wb[sheet_name]

            # Parse range boundaries
            parts = cell_range.upper().split(":")
            min_cell = parts[0]
            max_cell = parts[1] if len(parts) > 1 else parts[0]
            min_col_str, min_row = parse_cell_ref(min_cell)
            max_col_str, max_row = parse_cell_ref(max_cell)

            from openpyxl.utils import column_index_from_string

            min_col = column_index_from_string(min_col_str)
            max_col = column_index_from_string(max_col_str)

            # Build chart
            ctype = chart_type.lower()
            if ctype in ("bar", "column"):
                chart = BarChart()
                chart.type = "col" if ctype == "column" else "bar"
                chart.grouping = "clustered"
            elif ctype == "line":
                chart = LineChart()
            elif ctype == "pie":
                chart = PieChart()
            elif ctype == "scatter":
                chart = ScatterChart()
            elif ctype == "area":
                chart = AreaChart()
            else:
                chart = BarChart()
                chart.type = "col"

            chart.title = title
            chart.style = 10
            chart.width = 15
            chart.height = 10

            data = Reference(
                ws_data, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
            )
            chart.add_data(data, titles_from_data=True)

            # Target sheet
            if target_sheet:
                if target_sheet not in wb.sheetnames:
                    wb.create_sheet(target_sheet)
                ws_target = wb[target_sheet]
            else:
                ws_target = ws_data

            # Auto-position: place chart to the right of data
            anchor_col = idx_to_col(max_col)  # one column right of data
            anchor = f"{anchor_col}{min_row}"
            ws_target.add_chart(chart, anchor)

            save_workbook_safely(wb, path)
            log.info("excelmcp_create_chart OK: %s '%s' at %s", chart_type, title, anchor)
            return {
                "ok": True,
                "data": {
                    "chart_type": chart_type,
                    "title": title,
                    "positioned_at": anchor,
                    "sheet": ws_target.title,
                    "backend": "openpyxl",
                },
            }
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


def excelmcp_create_pivot_table(
    file_path: str,
    source_range: str,
    rows: list[str],
    columns: list[str],
    values: list[str],
    target_sheet: str | None = None,
) -> dict:
    """
    Create a pivot summary table (openpyxl doesn't support native PivotTables).
    Generates a SUMIFS-based aggregation table on target_sheet.
    """
    path = Path(file_path)
    if not path.exists():
        return _error(f"File not found: {file_path}")

    # Direct COM pivot summary creation (if Excel has file open)
    if excel_is_open(file_path):
        try:
            import win32com.client
            import pythoncom

            pythoncom.CoInitialize()

            target = Path(file_path).resolve()
            xl = None
            wb = None
            try:
                wb = win32com.client.GetObject(str(target))
                xl = wb.Application
            except Exception:
                xl = win32com.client.GetActiveObject("Excel.Application")
                for w in xl.Workbooks:
                    if Path(w.FullName).resolve() == target:
                        wb = w
                        break

            if wb is not None:
                sheet_name, cell_range = _parse_range_spec(source_range, wb.ActiveSheet.Name)
                ws_src = wb.Sheets(sheet_name)

                parts = cell_range.upper().split(":")
                min_cell = parts[0]
                max_cell = parts[1] if len(parts) > 1 else parts[0]
                min_col_str, min_row = parse_cell_ref(min_cell)
                max_col_str, max_row = parse_cell_ref(max_cell)

                min_col = col_to_idx(min_col_str) + 1
                max_col = col_to_idx(max_col_str) + 1

                header_map = {}
                for c in range(min_col, max_col + 1):
                    h = ws_src.Cells(min_row, c).Value
                    if h:
                        header_map[str(h)] = idx_to_col(c - 1)

                tsheet = target_sheet or (sheet_name + "_Pivot")
                # Delete existing sheet if it exists
                xl.DisplayAlerts = False
                try:
                    wb.Sheets(tsheet).Delete()
                except Exception:
                    pass
                xl.DisplayAlerts = True

                ws_pivot = wb.Sheets.Add(Before=wb.Sheets(1))
                ws_pivot.Name = tsheet

                pivot_headers = rows + columns + values
                for i, h in enumerate(pivot_headers, start=1):
                    ws_pivot.Cells(1, i).Value = h

                formula_row = 2
                for val_col in values:
                    val_letter = header_map.get(val_col, "A")
                    for row_field in rows:
                        row_letter = header_map.get(row_field, "A")
                        ws_pivot.Cells(formula_row, 1).Value = f"={row_field} totals"
                        formula = (
                            f"=SUMIF({sheet_name}!{row_letter}:{row_letter},"
                            f'"<>",'
                            f"{sheet_name}!{val_letter}:{val_letter})"
                        )
                        ws_pivot.Cells(formula_row, len(pivot_headers)).Formula = formula
                        formula_row += 1

                # Save changes
                wb.Save()
                log.info("Direct win32com pivot summary creation OK → sheet '%s'", tsheet)
                return {
                    "ok": True,
                    "data": {
                        "method": "summary_table",
                        "sheet": tsheet,
                        "rows_fields": rows,
                        "columns_fields": columns,
                        "values_fields": values,
                        "backend": "win32com",
                    },
                }
        except Exception as e:
            log.warning(
                "Direct COM pivot table creation failed: %s. Falling back to openpyxl...", e
            )

    try:
        with excel_com_lease(file_path):
            import openpyxl
            from openpyxl.utils import column_index_from_string, get_column_letter

            wb = openpyxl.load_workbook(str(path))
            sheet_name, cell_range = _parse_range_spec(source_range, wb.active.title)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.active.title
            ws_src = wb[sheet_name]

            # Read header row to map column names → column letters
            header_map: dict[str, str] = {}
            parts = cell_range.upper().split(":")
            min_cell = parts[0]
            _, min_row = parse_cell_ref(min_cell)
            max_cell = parts[1] if len(parts) > 1 else parts[0]
            max_col_str, max_row = parse_cell_ref(max_cell)
            max_col = column_index_from_string(max_col_str)
            min_col_str, _ = parse_cell_ref(min_cell)
            min_col = column_index_from_string(min_col_str)

            for c in range(min_col, max_col + 1):
                h = ws_src.cell(row=min_row, column=c).value
                if h:
                    header_map[str(h)] = get_column_letter(c)

            # Create / clear target sheet
            tsheet = target_sheet or (sheet_name + "_Pivot")
            if tsheet in wb.sheetnames:
                del wb[tsheet]
            ws_pivot = wb.create_sheet(tsheet)

            # Write pivot header
            pivot_headers = rows + columns + values
            for i, h in enumerate(pivot_headers, start=1):
                ws_pivot.cell(row=1, column=i).value = h

            # Write a SUMIFS formula row for each value column
            formula_row = 2
            for val_col in values:
                val_letter = header_map.get(val_col, "A")
                for row_field in rows:
                    row_letter = header_map.get(row_field, "A")
                    ws_pivot.cell(row=formula_row, column=1).value = f"={row_field} totals"
                    # SUMIF formula: =SUMIF(Sheet1!A:A,"<>",Sheet1!C:C)
                    formula = (
                        f"=SUMIF({sheet_name}!{row_letter}:{row_letter},"
                        f'"<>",'
                        f"{sheet_name}!{val_letter}:{val_letter})"
                    )
                    ws_pivot.cell(row=formula_row, column=len(pivot_headers)).value = formula
                    formula_row += 1

            save_workbook_safely(wb, path)
            log.info("excelmcp_create_pivot_table OK → sheet '%s'", tsheet)
            return {
                "ok": True,
                "data": {
                    "method": "summary_table",
                    "sheet": tsheet,
                    "rows_fields": rows,
                    "columns_fields": columns,
                    "values_fields": values,
                    "backend": "openpyxl",
                },
            }
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


def excelmcp_screenshot_range(
    file_path: str, range_spec: str, output_path: str | None = None
) -> dict:
    """
    Capture a range as PNG. Requires win32com + Excel open OR Pillow + openpyxl.
    Returns error gracefully if unavailable (not a fatal failure).
    """
    # Try via ExcelMcp CLI if available
    if excelmcp_available():
        try:
            if output_path is None:
                tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                output_path = tmp.name
                tmp.close()
            args = {
                "file": file_path,
                "range": range_spec,
                "output": output_path,
            }
            result = subprocess.run(
                ["excel-mcp-server", "screenshot", json.dumps(args)],
                capture_output=True,
                text=True,
                timeout=20,
            )
            if result.returncode == 0 and Path(output_path).exists():
                return {"ok": True, "data": {"output_path": output_path}}
        except Exception as e:
            log.warning("ExcelMcp screenshot failed: %s", e)

    return {
        "ok": False,
        "data": None,
        "error": (
            "Screenshot requires excel-mcp-server or win32com with Excel open. "
            "Install with: pip install excel-mcp-server"
        ),
    }


# ──────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ──────────────────────────────────────────────────────────────────────────────


def _parse_range_spec(range_spec: str, default_sheet: str) -> tuple[str, str]:
    """Parse 'Sheet1!A1:D10' → ('Sheet1', 'A1:D10'). Handles missing sheet name."""
    if "!" in range_spec:
        sheet, cell_range = range_spec.split("!", 1)
        return sheet.strip("'\""), cell_range
    return default_sheet, range_spec


def _openpyxl_write_cell(
    file_path: str,
    cell: str,
    value: Any,
    formula: str | None,
    number_format: str | None,
    bold: bool,
) -> dict:
    """Write a single cell via openpyxl, preserving other cell formatting."""
    try:
        with excel_com_lease(file_path):
            import openpyxl

            wb = openpyxl.load_workbook(str(file_path), data_only=False)
            ws = wb.active
            target = ws[cell]

            # Write value or formula
            if formula:
                target.value = formula if formula.startswith("=") else f"={formula}"
            elif value is not None:
                target.value = value

            # Apply optional formatting (don't touch existing if not specified)
            if number_format:
                target.number_format = number_format
            if bold:
                from openpyxl.styles import Font

                existing_font = target.font
                target.font = Font(
                    bold=True,
                    name=existing_font.name if existing_font else "Calibri",
                    size=existing_font.size if existing_font else 11,
                    color=existing_font.color if existing_font else None,
                )

            save_workbook_safely(wb, file_path)
            log.info("_openpyxl_write_cell OK: %s = %s", cell, formula or value)
            return {
                "ok": True,
                "data": {
                    "cell": cell,
                    "written": formula or str(value),
                    "backend": "openpyxl",
                },
            }
    except ImportError:
        return _error("openpyxl not installed")
    except Exception:
        return _error(traceback.format_exc())


def _win32com_write_cell(file_path: str, cell: str, value: Any, formula: str | None) -> dict:
    """Write to a live Excel file via win32com COM automation."""
    try:
        import win32com.client
        import pythoncom

        pythoncom.CoInitialize()

        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
            excel_running = True
        except Exception:
            excel_running = False

        if excel_running:
            target_path = str(Path(file_path).resolve())
            try:
                wb = win32com.client.GetObject(target_path)
                ws = wb.ActiveSheet
                ws.Range(cell).Formula = formula if formula else (value or "")
                return {"ok": True, "data": {"cell": cell, "backend": "win32com_live"}}
            except Exception:
                try:
                    for wb in xl.Workbooks:
                        if str(Path(wb.FullName).resolve()) == target_path:
                            ws = wb.ActiveSheet
                            ws.Range(cell).Formula = formula if formula else (value or "")
                            return {"ok": True, "data": {"cell": cell, "backend": "win32com_live"}}
                except Exception:
                    pass
        return _error("Excel is not running or workbook is not open")
    except Exception as e:
        return _error(f"win32com write failed: {e}")


def _adjust_formula(formula: str, row_offset: int, col_offset: int) -> str:
    """
    Adjust relative cell references in a formula by row/col offset.
    Absolute references ($A$1) are left unchanged.
    e.g. '=A2-B2' with row_offset=1 → '=A3-B3'
    """
    if not formula.startswith("="):
        return formula

    def _adjust_ref(m: re.Match) -> str:
        col_abs = "$" in m.group(1)
        col_str = m.group(1).replace("$", "").upper()
        row_abs = "$" in m.group(2)
        row_num = int(m.group(2).replace("$", ""))

        new_col = col_str if col_abs else idx_to_col(col_to_idx(col_str) + col_offset)
        new_row = row_num if row_abs else row_num + row_offset

        col_part = ("$" + new_col) if col_abs else new_col
        row_part = ("$" + str(new_row)) if row_abs else str(max(1, new_row))
        return col_part + row_part

    # Match cell references: optional $, column letters, optional $, row digits
    pattern = r"(\$?[A-Za-z]+)(\$?\d+)"
    return re.sub(pattern, _adjust_ref, formula)


def _error(msg: str) -> dict:
    log.error("ExcelMcpBridge error: %s", msg[:200])
    return {"ok": False, "data": None, "error": msg}
