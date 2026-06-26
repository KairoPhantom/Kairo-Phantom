"""
Excel SmartContextCapture — Pi-inspired auto-context injection for Kairo Phantom.
==================================================================================
Captures rich workbook context before every Alt+Ctrl+M press in Excel.
The user never needs to describe what they're looking at.

Pattern adopted from tmustier/pi-for-excel:
- Workbook blueprint (sheets, tables, named ranges, charts)
- Active cell + surrounding grid with values AND formulas
- Column headers detection
- Named ranges for formula context
"""

from __future__ import annotations

import logging
import traceback
from pathlib import Path

log = logging.getLogger("kairo-sidecar.excel_context")


# ──────────────────────────────────────────────────────────────────────────────
# Core context capture class
# ──────────────────────────────────────────────────────────────────────────────


class ExcelContextCapture:
    """
    Captures rich spreadsheet context before every LLM call in Excel mode.
    Implements Pi's auto-context injection pattern.
    """

    def capture(self, file_path: str, active_cell: str | None = None) -> dict:
        """
        Capture full Excel context for the LLM.
        Returns a structured context dict ready for system prompt injection.

        Args:
            file_path: path to .xlsx file
            active_cell: e.g. "C7" (from UIA window title or last cursor position)

        Returns dict with keys: blueprint, active_cell, sheet_name, grid,
                                headers, named_ranges, tables, sheet_names, ...
        """
        path = Path(file_path)
        if not path.exists():
            log.warning("ExcelContextCapture: file not found: %s", file_path)
            return _empty_context(active_cell)

        active_sheet = None
        try:
            import win32com.client
            import pythoncom
            import subprocess

            pythoncom.CoInitialize()

            excel_running = False
            try:
                out = subprocess.run(
                    ["tasklist", "/FI", "IMAGENAME eq excel.exe"], capture_output=True, text=True
                )
                excel_running = "excel.exe" in out.stdout.lower()
            except Exception:
                pass

            if excel_running:
                target_path = str(Path(file_path).resolve())
                wb_com = None
                xl = None
                try:
                    wb_com = win32com.client.GetObject(target_path)
                    xl = wb_com.Application
                    active_sheet = wb_com.ActiveSheet.Name
                    active_cell = xl.ActiveCell.Address.replace("$", "")
                    log.info(
                        f"Excel COM detected active sheet via moniker: {active_sheet}, cell: {active_cell}"
                    )
                except Exception:
                    try:
                        xl = win32com.client.GetActiveObject("Excel.Application")
                        for wb in xl.Workbooks:
                            if str(Path(wb.FullName).resolve()) == target_path:
                                wb_com = wb
                                active_sheet = wb_com.ActiveSheet.Name
                                active_cell = xl.ActiveCell.Address.replace("$", "")
                                log.info(
                                    f"Excel COM detected active sheet via fallback: {active_sheet}, cell: {active_cell}"
                                )
                                break
                    except Exception:
                        pass
        except Exception as e:
            log.debug(f"Could not get Excel active info via COM: {e}")

        try:
            import openpyxl
            from openpyxl.utils import column_index_from_string, get_column_letter

            wb = openpyxl.load_workbook(str(path), data_only=False)
            if active_sheet and active_sheet in wb.sheetnames:
                ws = wb[active_sheet]
            else:
                ws = wb.active

            # Parse active cell
            if active_cell:
                col_str = "".join(c for c in active_cell if c.isalpha())
                row_num = int("".join(c for c in active_cell if c.isdigit()) or "1")
                col_num = column_index_from_string(col_str) if col_str else 1
            else:
                row_num, col_num = 1, 1
                active_cell = "A1"

            # Extract surrounding grid (±8 rows/cols)
            radius = 8
            r_start = max(1, row_num - radius)
            r_end = min(ws.max_row or row_num + radius, row_num + radius)
            c_start = max(1, col_num - radius)
            c_end = min(ws.max_column or col_num + radius, col_num + radius)

            grid: list[list[dict]] = []
            for r in range(r_start, r_end + 1):
                row_data = []
                for c in range(c_start, c_end + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    formula = ""
                    value_str = ""
                    if isinstance(val, str) and val.startswith("="):
                        formula = val
                    elif val is not None:
                        value_str = str(val)
                    row_data.append(
                        {
                            "ref": cell.coordinate,
                            "value": value_str,
                            "formula": formula,
                            "is_active": (r == row_num and c == col_num),
                            "number_format": cell.number_format or "General",
                        }
                    )
                grid.append(row_data)

            # Column headers (row 1 values)
            headers: dict[str, str] = {}
            for c in range(1, (ws.max_column or 1) + 1):
                val = ws.cell(row=1, column=c).value
                if val:
                    headers[get_column_letter(c)] = str(val)

            # Named ranges
            try:
                named_ranges = list(wb.defined_names.keys())
            except Exception:
                named_ranges = []

            # Tables per sheet
            tables: list[dict] = []
            try:
                for sname in wb.sheetnames:
                    ws2 = wb[sname]
                    for tname, tobj in (getattr(ws2, "tables", {}) or {}).items():
                        tables.append(
                            {"name": tname, "sheet": sname, "range": getattr(tobj, "ref", "")}
                        )
            except Exception:
                pass

            # Workbook blueprint
            blueprint = {
                "sheets": [
                    {
                        "name": n,
                        "max_row": wb[n].max_row or 0,
                        "max_col": wb[n].max_column or 0,
                    }
                    for n in wb.sheetnames
                ],
                "named_ranges": named_ranges,
                "tables": tables,
                "total_sheets": len(wb.sheetnames),
            }

            return {
                "blueprint": blueprint,
                "active_cell": active_cell,
                "sheet_name": ws.title,
                "sheet_names": wb.sheetnames,
                "grid": grid,
                "headers": headers,
                "named_ranges": named_ranges,
                "tables": tables,
                "max_row": ws.max_row or 0,
                "max_col": ws.max_column or 0,
                "file_path": str(path),
            }
        except ImportError:
            log.error("openpyxl not installed")
            return _empty_context(active_cell)
        except Exception:
            log.error("ExcelContextCapture.capture error: %s", traceback.format_exc())
            return _empty_context(active_cell)

    def to_system_prompt_fragment(self, context: dict) -> str:
        """
        Convert captured context to a system prompt fragment for the LLM.
        Tells the AI exactly what spreadsheet it's working with.
        """
        lines: list[str] = []

        lines.append("## Excel Context")

        # Sheet info
        bp = context.get("blueprint", {})
        sheet_names = context.get("sheet_names") or [s["name"] for s in bp.get("sheets", [])]
        if sheet_names:
            lines.append(
                f"Workbook has {len(sheet_names)} sheet(s): {', '.join(str(s) for s in sheet_names)}"
            )
        lines.append(
            f"Active sheet: {context.get('sheet_name', 'Sheet1')} | Active cell: {context.get('active_cell', 'A1')}"
        )

        # Headers
        headers = context.get("headers", {})
        if headers:
            h_str = ", ".join(f"{col}={name}" for col, name in list(headers.items())[:12])
            lines.append(f"Column headers: {h_str}")

        # Named ranges
        named = context.get("named_ranges", [])
        if named:
            lines.append(f"Named ranges: {', '.join(str(n) for n in named[:10])}")

        # Tables
        tables = context.get("tables", [])
        if tables:
            t_str = ", ".join(t["name"] for t in tables[:5])
            lines.append(f"Tables: {t_str}")

        # Surrounding data grid (compact representation)
        grid = context.get("grid", [])
        if grid:
            lines.append(
                f"\n## Surrounding Data (±8 rows/cols from {context.get('active_cell', 'A1')})"
            )
            for row in grid[:16]:  # cap at 16 rows
                row_strs = []
                for cell in row[:16]:  # cap at 16 cols
                    content = cell.get("formula") or cell.get("value") or ""
                    prefix = "►" if cell.get("is_active") else ""
                    row_strs.append(f"{prefix}[{cell['ref']}]={content[:20]}")
                lines.append(" | ".join(row_strs))

        # Instructions
        lines.append("\n## Instructions for Excel Mode")
        lines.append("- Output ONLY valid JSON: array of ExcelWriteOp objects")
        lines.append(
            '- ExcelWriteOp fields: {"cell": "F2", "formula": "=C2-B2", "value": "", "number_format": "0.00%"}'
        )
        lines.append("- Use = prefix for ALL formulas (e.g. =SUM(A1:A10))")
        lines.append('- Use commas as separators in formulas (=SUMIF(A:A,"West",B:B))')
        lines.append("- Use named ranges where available instead of raw cell references")
        lines.append(
            '- For charts: {"chart_type": "line", "source_range": "A1:B100", "title": "Revenue"}'
        )
        lines.append(
            '- For pivots: {"source_range": "A1:D100", "rows": ["Region"], "columns": [], "values": ["Revenue"]}'
        )

        return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
# Standalone functions (used by sidecar dispatcher)
# ──────────────────────────────────────────────────────────────────────────────


def get_workbook_overview(file_path: str) -> dict:
    """
    Get structural overview: sheets, tables, charts, pivots, named ranges.
    Returns: {sheets, named_ranges, tables, total_sheets, total_named_ranges}
    """
    from sidecar.parsers.excelmcp_bridge import get_workbook_blueprint

    result = get_workbook_blueprint(file_path)
    if not result.get("ok"):
        return result
    return result


def get_active_cell_context(file_path: str, active_cell: str, radius: int = 8) -> dict:
    """
    Read surrounding grid around the active cell.
    Returns: {active_cell, grid, headers, sheet_name}
    """
    cap = ExcelContextCapture()
    ctx = cap.capture(file_path, active_cell)
    # Trim grid to requested radius
    grid = ctx.get("grid", [])
    trimmed = [row[: radius * 2 + 1] for row in grid[: radius * 2 + 1]]
    return {
        "ok": True,
        "data": {
            "active_cell": ctx.get("active_cell"),
            "sheet_name": ctx.get("sheet_name"),
            "grid": trimmed,
            "headers": ctx.get("headers", {}),
        },
    }


def format_excel_context_for_prompt(file_path: str, active_cell: str, command: str) -> str:
    """
    Full pipeline: capture context + format as system prompt fragment + append command.
    Returns the complete string ready to prepend to the LLM system prompt.
    """
    cap = ExcelContextCapture()
    try:
        ctx = cap.capture(file_path, active_cell)
        fragment = cap.to_system_prompt_fragment(ctx)
    except Exception:
        fragment = (
            "## Excel Context\n"
            f"Active cell: {active_cell}\n"
            "(Context capture unavailable — openpyxl may need to be installed)\n"
        )
    return f"{fragment}\n\n## User Command\n{command}"


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────


def _empty_context(active_cell: str | None) -> dict:
    return {
        "blueprint": {"sheets": [], "named_ranges": [], "tables": [], "total_sheets": 0},
        "active_cell": active_cell or "A1",
        "sheet_name": "Sheet1",
        "sheet_names": ["Sheet1"],
        "grid": [],
        "headers": {},
        "named_ranges": [],
        "tables": [],
        "max_row": 0,
        "max_col": 0,
        "file_path": "",
    }
