"""
sidecar/creators/xlsx_creator.py — Kairo Phantom Create-From-Scratch XLSX
==========================================================================
Creates a new .xlsx workbook from a structured content dict and opens it in Excel.

Usage
-----
    from sidecar.creators.xlsx_creator import XlsxCreator

    creator = XlsxCreator()
    path = creator.create_and_open({
        "title": "Q3 Revenue Tracker",
        "author": "Kairo Phantom",
        "sheets": [
            {
                "name": "Summary",
                "headers": ["Product", "Revenue", "Cost", "Margin"],
                "rows": [
                    ["Widget A", 10000, 6000, "=IFERROR((B2-C2)/B2,0)"],
                    ["Widget B", 8000, 5000, "=IFERROR((B3-C3)/B3,0)"]
                ],
                "totals": True
            },
            {
                "name": "Notes",
                "cells": [
                    {"cell": "A1", "value": "Prepared by Kairo Phantom"},
                    {"cell": "A2", "formula": "=TODAY()"}
                ]
            }
        ]
    })
"""

import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional

log = logging.getLogger("kairo-sidecar.xlsx_creator")

KAIRO_DOCS_DIR = Path.home() / "Documents" / "Kairo"


class XlsxCreator:
    """
    Creates new .xlsx workbooks from a structured content dict.
    Saves to ~/Documents/Kairo/ and opens in Excel via os.startfile().
    """

    def create(
        self,
        content: Dict[str, Any],
        output_path: Optional[str] = None,
    ) -> str:
        """
        Creates a .xlsx from structured content.

        Parameters
        ----------
        content     : Dict with keys:
                      - title (str): Workbook title (core property)
                      - author (str): Author name (core property)
                      - sheets (list): Each sheet dict may have:
                          - name (str): Sheet tab name
                          - headers (list[str]): Column headers in row 1
                          - rows (list[list]): Data rows (may include formulas)
                          - totals (bool): If True, adds SUM formula row
                          - cells (list[dict]): Ad-hoc cell writes:
                              {"cell": "A1", "value": "text"}
                              {"cell": "B2", "formula": "=SUM(B2:B10)"}
        output_path : Full path to save. Auto-generated if None.

        Returns
        -------
        str — absolute path of the created file.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        # Remove default empty sheet if we'll add our own
        default_sheet = wb.active

        title = content.get("title", "Untitled Workbook")
        sheets_data = content.get("sheets", [])

        if not sheets_data:
            # Just set default sheet name to title
            default_sheet.title = title[:31]
        else:
            # Remove default empty sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            for sheet_data in sheets_data:
                sheet_name = str(sheet_data.get("name", "Sheet"))[:31]
                ws = wb.create_sheet(title=sheet_name)

                headers = sheet_data.get("headers", [])
                rows_data = sheet_data.get("rows", [])
                add_totals = sheet_data.get("totals", False)

                # Write headers in row 1 with bold + fill
                if headers:
                    header_fill = PatternFill("solid", fgColor="366092")
                    header_font = Font(bold=True, color="FFFFFF")
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=str(header))
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                # Write data rows
                data_start_row = 2 if headers else 1
                for row_idx, row_data in enumerate(rows_data, start=data_start_row):
                    for col_idx, cell_val in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell_val, str) and cell_val.startswith("="):
                            cell.value = cell_val  # formula
                        else:
                            cell.value = cell_val

                # Add totals row if requested
                if add_totals and headers and rows_data:
                    totals_row = data_start_row + len(rows_data)
                    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
                    for col_idx in range(2, len(headers) + 1):
                        col_letter = get_column_letter(col_idx)
                        end_row = totals_row - 1
                        ws.cell(
                            row=totals_row,
                            column=col_idx,
                            value=f'=IFERROR(SUM({col_letter}{data_start_row}:{col_letter}{end_row}),"")',
                        ).font = Font(bold=True)

                # Ad-hoc cell writes
                for cell_write in sheet_data.get("cells", []):
                    addr = cell_write.get("cell", "A1")
                    if "formula" in cell_write:
                        ws[addr] = cell_write["formula"]
                    elif "value" in cell_write:
                        ws[addr] = cell_write["value"]

                # Auto-fit column widths
                for column_cells in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(column_cells[0].column)
                    for cell in column_cells:
                        try:
                            cell_len = len(str(cell.value)) if cell.value is not None else 0
                            if cell_len > max_length:
                                max_length = cell_len
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = min(50, max(8, max_length + 2))

        # Set core properties
        try:
            wb.properties.title = title
            wb.properties.creator = content.get("author", "Kairo Phantom")
        except Exception:
            pass

        # Resolve output path
        if not output_path:
            KAIRO_DOCS_DIR.mkdir(parents=True, exist_ok=True)
            safe_title = (
                "".join(c if c.isalnum() or c in " _-" else "_" for c in title).strip()[:50]
                or "Workbook"
            )
            output_path = str(KAIRO_DOCS_DIR / f"{safe_title}.xlsx")

        wb.save(output_path)
        log.info(f"XlsxCreator: saved to {output_path}")
        return output_path

    def create_and_open(self, content: Dict[str, Any]) -> str:
        """Creates a .xlsx and opens it in Excel."""
        path = self.create(content)
        try:
            os.startfile(path)
            log.info(f"XlsxCreator: opened {path}")
        except Exception as e:
            log.warning(f"XlsxCreator: could not auto-open {path}: {e}")
        return path
