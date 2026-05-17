#!/usr/bin/env python3
"""
Quick test for the Kairo sidecar.
Usage: python kairo-sidecar/test_sidecar.py
"""
import socket
import json
import time
import sys

HOST = "127.0.0.1"
PORT = 7438


def send_request(action: str, path: str = "", payload: dict = None) -> dict:
    sock = socket.create_connection((HOST, PORT), timeout=5)
    req = json.dumps({"id": f"test-{action}", "action": action, "path": path, "payload": payload or {}})
    sock.sendall((req + "\n").encode())
    data = b""
    while not data.endswith(b"\n"):
        chunk = sock.recv(4096)
        if not chunk:
            break
        data += chunk
    sock.close()
    return json.loads(data.decode().strip())


def main():
    print("=" * 50)
    print("Kairo Sidecar Test Suite")
    print("=" * 50)
    
    # 1. Ping
    try:
        r = send_request("ping")
        assert r["ok"], f"Ping failed: {r}"
        print(f"✅ Ping: {r['data']}")
    except Exception as e:
        print(f"❌ Sidecar not running on {HOST}:{PORT}: {e}")
        print("   Run: python kairo-sidecar/sidecar.py")
        sys.exit(1)

    # 2. Test DOCX read (use planning test doc if exists)
    import os
    test_docx = os.path.join(os.path.dirname(__file__), ".planning", "Kairo-test.docx")
    if os.path.exists(test_docx):
        r = send_request("read_docx", path=test_docx)
        if r["ok"]:
            d = r["data"]
            print(f"✅ DOCX read: {d.get('paragraph_count', 0)} paragraphs, {len(d.get('headings', []))} headings")
            print(f"   Headings: {[h['text'] for h in d.get('headings', [])[:3]]}")
        else:
            print(f"⚠️  DOCX read error: {r.get('error')}")
    else:
        print(f"⚠️  No test .docx at {test_docx} — skipping DOCX test")

    # 3. Test XLSX read (create a temp xlsx)
    try:
        import openpyxl, tempfile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Product"
        ws["B1"] = "Sales"
        ws["A2"] = "Widget"
        ws["B2"] = 1500
        ws["A3"] = "Gadget"
        ws["B3"] = 2300
        tf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tf.name)
        tf.close()

        r = send_request("read_xlsx", path=tf.name, payload={"active_cell": "B2"})
        if r["ok"]:
            d = r["data"]
            print(f"✅ XLSX read: active={d.get('active_cell')} sheet={d.get('sheet_name')}")
            print(f"   Headers: {d.get('headers')}")
        else:
            print(f"⚠️  XLSX read error: {r.get('error')}")

        # Write a formula
        r2 = send_request("write_xlsx", path=tf.name, payload={
            "operations": [{"cell": "C2", "formula": "=B2*0.05", "value": ""}]
        })
        if r2["ok"]:
            print(f"✅ XLSX write: formula written to C2")
        else:
            print(f"⚠️  XLSX write error: {r2.get('error')}")

        os.unlink(tf.name)
    except ImportError:
        print("⚠️  openpyxl not installed — skipping Excel test")

    print("=" * 50)
    print("✅ Sidecar tests complete!")


if __name__ == "__main__":
    main()
