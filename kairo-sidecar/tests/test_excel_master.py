import os
import sys
import tempfile
import time
import shutil
import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

# Add sidecar package to path
sys.path.insert(0, str(Path(__file__).parent.parent.resolve()))

from sidecar.masters.excel_master import (
    ExcelContext,
    ExcelContextExtractor,
    ExcelOperationValidator,
    ExcelWriter,
    ValidationResult,
    ForgeValidator
)

@pytest.fixture
def temp_xlsx():
    """Create a basic Excel workbook with some values, formulas, named ranges, and conditional formatting."""
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, "test.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    ws["A1"] = "Product"
    ws["B1"] = "Price"
    ws["C1"] = "Quantity"
    ws["D1"] = "Total"
    
    ws["A2"] = "Widget A"
    ws["B2"] = 10.0
    ws["C2"] = 5
    ws["D2"] = "=B2*C2"
    
    # Add named range
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("PriceRange", attr_text="Sales!$B$2:$B$2"))
    
    # Add conditional formatting
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    rule = CellIsRule(operator="greaterThan", formula=["5"], fill=fill)
    ws.conditional_formatting.add("B2", rule)
    
    wb.save(file_path)
    yield file_path
    shutil.rmtree(temp_dir)


# --- Test 1: SUM formula written correctly, evaluates on file open ---
def test_scenario_1_sum_formula_evaluates_on_open(temp_xlsx):
    writer = ExcelWriter()
    op = {"type": "write_cell", "cell": "D3", "formula": "=SUM(B2:C2)"}
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) == 0
        
        wb = openpyxl.load_workbook(temp_xlsx, data_only=False)
        assert wb.active["D3"].value == "=SUM(B2:C2)"


# --- Test 2: Locale auto-fix (EU semicolons) ---
def test_scenario_2_locale_semicolon_auto_fix():
    validator = ExcelOperationValidator()
    
    # Semicolons corrected to commas in English locale
    context_en = MagicMock()
    context_en.active_cell = "E2"
    context_en.locale = "en"
    op_en = {"type": "write_cell", "cell": "E2", "formula": "=SUM(B2;C2)"}
    res_en = validator.validate(op_en, context_en)
    assert res_en.valid is True
    assert op_en["formula"] == "=SUM(B2,C2)"
    
    # Commas corrected to semicolons in EU locale
    context_eu = MagicMock()
    context_eu.active_cell = "E2"
    context_eu.locale = "eu"
    op_eu = {"type": "write_cell", "cell": "E2", "formula": "=SUM(B2,C2)"}
    res_eu = validator.validate(op_eu, context_eu)
    assert res_eu.valid is True
    assert op_eu["formula"] == "=SUM(B2;C2)"


# --- Test 3: VLOOKUP missing FALSE → auto-corrected to include FALSE ---
def test_scenario_3_vlookup_missing_false_corrected():
    validator = ExcelOperationValidator()
    context = MagicMock()
    context.active_cell = "E2"
    context.locale = "en"
    
    # Test VLOOKUP with 3 args (missing FALSE)
    op = {"type": "write_cell", "cell": "E2", "formula": "=VLOOKUP(A2,B:D,3)"}
    res = validator.validate(op, context)
    assert res.valid is True
    assert op["formula"] == "=VLOOKUP(A2,B:D,3,FALSE)"
    
    # Test VLOOKUP with 2 args (missing col_idx and FALSE)
    op2 = {"type": "write_cell", "cell": "E2", "formula": "=VLOOKUP(A2,B:D)"}
    res2 = validator.validate(op2, context)
    assert res2.valid is True
    assert op2["formula"] == "=VLOOKUP(A2,B:D,2,FALSE)"


# --- Test 4: Adjacent cells completely unchanged (binary comparison per-cell) ---
def test_scenario_4_adjacent_cells_unchanged(temp_xlsx):
    writer = ExcelWriter()
    op = {"type": "write_cell", "cell": "C2", "value": 99.0}
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        # Apply operation
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) == 0
        
        # Load and assert
        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws["C2"].value == 99.0  # Modified cell
        assert ws["B2"].value == 10.0  # Adjacent untouched
        assert ws["D2"].value == "=B2*C2"  # Adjacent untouched


# --- Test 5: Named range used in formula → resolves correctly ---
def test_scenario_5_named_range_resolves_in_formula(temp_xlsx):
    extractor = ExcelContextExtractor()
    ctx = extractor.extract(temp_xlsx, "D3", "Sales")
    
    # PriceRange must be present as a key in named_ranges dict
    assert "PriceRange" in ctx.named_ranges
    
    validator = ExcelOperationValidator()
    op = {"type": "write_cell", "cell": "D3", "formula": "=SUM(PriceRange)"}
    res = validator.validate(op, ctx)
    assert res.valid is True


# --- Test 6: Formula referencing another sheet → correct syntax ---
def test_scenario_6_cross_sheet_formula_syntax(temp_xlsx):
    # Add another sheet first
    wb = openpyxl.load_workbook(temp_xlsx)
    wb.create_sheet("Summary")
    wb.save(temp_xlsx)
    
    writer = ExcelWriter()
    op = {"type": "write_cell", "sheet": "Summary", "cell": "A1", "formula": "=Sales!B2*2"}
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) == 0
        
        wb_check = openpyxl.load_workbook(temp_xlsx)
        assert wb_check["Summary"]["A1"].value == "=Sales!B2*2"


# --- Test 7: Protected sheet → graceful error + formula shown in GRP for manual copy ---
def test_scenario_7_protected_sheet_graceful_error(temp_xlsx):
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb.active
    ws.protection.sheet = True # enable sheet protection
    wb.save(temp_xlsx)
    
    writer = ExcelWriter()
    op = {"type": "write_cell", "cell": "C2", "formula": "=SUM(A2:B2)"}
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) > 0
        assert "is protected" in res["errors"][0]
        assert "=SUM(A2:B2)" in res["errors"][0] # Manual copy formula shown in error


# --- Test 8: Invalid function name → validator rejects before write ---
def test_scenario_8_invalid_function_rejected():
    validator = ExcelOperationValidator()
    context = MagicMock()
    context.active_cell = "E2"
    context.locale = "en"
    
    op = {"type": "write_cell", "cell": "E2", "formula": "=SUPER_SUM(B2:C2)"}
    res = validator.validate(op, context)
    assert res.valid is False
    assert "Unknown Excel functions" in res.error


# --- Test 9: 10,000 row spreadsheet → context extraction under 3s ---
def test_scenario_9_large_spreadsheet_performance():
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, "large.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LargeData"
        # Fast append
        ws.append(("Col A", "Col B"))
        for r in range(2, 10002):
            ws.append((f"Val{r}", r))
        wb.save(file_path)
        
        extractor = ExcelContextExtractor()
        start = time.time()
        ctx = extractor.extract(file_path, "B5000", "LargeData")
        elapsed = time.time() - start
        
        assert elapsed < 8.0, f"Large spreadsheet context extraction took {elapsed}s"
        assert ctx.max_row >= 10000
        assert len(ctx.cells) > 0
    finally:
        shutil.rmtree(temp_dir)


# --- Test 10: Unbalanced parentheses → auto-fix attempted, shown in GRP ---
def test_scenario_10_unbalanced_parentheses_autofix():
    validator = ExcelOperationValidator()
    context = MagicMock()
    context.active_cell = "E2"
    context.locale = "en"
    
    # Auto-fix balances parentheses
    op = {"type": "write_cell", "cell": "E2", "formula": "=SUM(B2:C2"}
    res = validator.validate(op, context)
    assert res.valid is True
    assert op["formula"] == "=SUM(B2:C2)"


# --- Test 11: Write range B1:E1 → only those 4 cells changed ---
def test_scenario_11_write_range_boundary(temp_xlsx):
    writer = ExcelWriter()
    op = {
        "type": "write_range",
        "start_cell": "B3",
        "values": [[100, 200, 300, 400]]
    }
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) == 0
        
        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        
        # B3, C3, D3, E3 are modified
        assert ws["B3"].value == 100
        assert ws["C3"].value == 200
        assert ws["D3"].value == 300
        assert ws["E3"].value == 400
        
        # Adjacent cells A3 and F3 remain empty/unchanged
        assert ws["A3"].value is None
        assert ws["F3"].value is None


# --- Test 12: keep_vba=True → macros preserved after openpyxl write ---
@patch("sidecar.masters.excel_master.load_workbook")
def test_scenario_12_keep_vba_macros_preserved(mock_load):
    mock_wb = MagicMock()
    mock_ws = MagicMock()
    mock_ws.protection.sheet = False
    mock_ws.protection.enabled = False
    mock_wb.sheetnames = ["Sales"]
    mock_wb.__getitem__.return_value = mock_ws
    mock_wb.active = mock_ws
    mock_load.return_value = mock_wb
    
    writer = ExcelWriter()
    writer.apply_operations("dummy.xlsx", [{"type": "write_cell", "cell": "A1", "value": 42}])
    
    # Assert load_workbook was called with keep_vba=True
    mock_load.assert_called_with("dummy.xlsx", keep_vba=True)


# --- Test 13: Atomic write → kill sidecar mid-save → original intact ---
def test_scenario_13_atomic_write_kill_mid_save(temp_xlsx):
    writer = ExcelWriter()
    op = {"type": "write_cell", "cell": "B2", "value": 999.0}
    
    original_size = os.path.getsize(temp_xlsx)
    
    with patch("openpyxl.workbook.workbook.Workbook.save", side_effect=IOError("Simulated write interruption")):
        with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
            res = writer.apply_operations(temp_xlsx, [op])
            assert len(res["errors"]) > 0
            
            # Original file remains completely unchanged
            assert os.path.exists(temp_xlsx)
            assert os.path.getsize(temp_xlsx) == original_size
            wb = openpyxl.load_workbook(temp_xlsx)
            assert wb.active["B2"].value == 10.0 # Remains original value


# --- Test 14: Number format preserved on cell after formula write ---
def test_scenario_14_number_format_preserved_after_write(temp_xlsx):
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb.active
    ws["B2"].number_format = "$#,##0.00"
    wb.save(temp_xlsx)
    
    writer = ExcelWriter()
    op = {"type": "write_cell", "cell": "B2", "formula": "=10+5"}
    
    with patch("sidecar.parsers.excelmcp_bridge.excel_is_open", return_value=False):
        res = writer.apply_operations(temp_xlsx, [op])
        assert len(res["errors"]) == 0
        
        wb_check = openpyxl.load_workbook(temp_xlsx)
        cell = wb_check.active["B2"]
        assert cell.value == "=10+5"
        assert cell.number_format == "$#,##0.00" # Format preserved


# --- Test 15: Header row detection correct on 5 different spreadsheet layouts ---
def test_scenario_15_header_row_detection_layouts():
    extractor = ExcelContextExtractor()
    
    def get_extracted_headers(setup_func):
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, "layout.xlsx")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            setup_func(ws)
            wb.save(file_path)
            ctx = extractor.extract(file_path, "A1")
            return ctx.headers
        finally:
            shutil.rmtree(temp_dir)
            
    # Layout 1: Standard row 1
    def layout1(ws):
        ws.append(("Date", "Product", "Revenue"))
        ws.append(("2026-01-01", "Widget A", 100))
    headers1 = get_extracted_headers(layout1)
    assert headers1["A"] == "Date"
    assert headers1["B"] == "Product"
    assert headers1["C"] == "Revenue"
    
    # Layout 2: Empty rows at top
    def layout2(ws):
        ws.cell(row=3, column=1, value="ID")
        ws.cell(row=3, column=2, value="Name")
        ws.cell(row=4, column=1, value=1)
        ws.cell(row=4, column=2, value="Item A")
    headers2 = get_extracted_headers(layout2)
    assert headers2["A"] == "ID"
    assert headers2["B"] == "Name"
    
    # Layout 3: Title row
    def layout3(ws):
        ws["A1"] = "Q1 2026 Sales Report"
        ws.cell(row=3, column=1, value="Category")
        ws.cell(row=3, column=2, value="Total Sales")
        ws.cell(row=4, column=1, value="Hardware")
        ws.cell(row=4, column=2, value=5000)
    headers3 = get_extracted_headers(layout3)
    assert headers3["A"] == "Category"
    assert headers3["B"] == "Total Sales"
    
    # Layout 4: Partially filled
    def layout4(ws):
        ws["A1"] = "Month"
        ws["B1"] = None
        ws["C1"] = "Expense"
        ws.append(("Jan", 100, 200))
    headers4 = get_extracted_headers(layout4)
    assert headers4["A"] == "Month"
    assert "B" not in headers4
    assert headers4["C"] == "Expense"
    
    # Layout 5: Numeric data in row 1, header on row 2
    def layout5(ws):
        ws["A1"] = 12345
        ws["B1"] = 67890
        ws.cell(row=2, column=1, value="Region")
        ws.cell(row=2, column=2, value="Manager")
        ws.cell(row=3, column=1, value="North")
        ws.cell(row=3, column=2, value="Alice")
    headers5 = get_extracted_headers(layout5)
    assert headers5["A"] == "Region"
    assert headers5["B"] == "Manager"
